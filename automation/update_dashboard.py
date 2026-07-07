"""
KFC Sales Command Centre — Daily Auto-Update
=============================================
Διαβάζει το πιο πρόσφατο Sales Analysis email από Outlook,
παράγει νέο data.json, και το κάνει push στο GitHub repo.

Τρέχει σιωπηλά (χωρίς GUI). Όλα τα logs γράφονται στο update.log.

Χρήση:
  python update_dashboard.py

Εξαρτήσεις:
  pip install pywin32 openpyxl pandas python-calamine
"""

import os
import sys
import json
import shutil
import subprocess
import logging
from datetime import datetime, timedelta
from pathlib import Path

# ============================================================
# CONFIGURATION
# ============================================================

# =====================================================================
# NOTIFICATION
# =====================================================================
def send_failure_notification(reason, details=""):
    """Send failure notification email via Outlook to NOTIFY_TO."""
    NOTIFY_TO = "report@frt-ike.gr"
    try:
        import win32com.client
        outlook_app = win32com.client.Dispatch("Outlook.Application")
        mail = outlook_app.CreateItem(0)
        mail.To = NOTIFY_TO
        from datetime import datetime as _dt_now
        mail.Subject = f"[STRICT MODE FAIL] KFC Sales Pipeline — {reason} {_dt_now().now().strftime('%Y-%m-%d %H:%M')}"
        mail.Body = (
            f"KFC Sales Pipeline notification\n"
            f"{'=' * 50}\n\n"
            f"Time:    {_dt_now().now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Reason:  {reason}\n\n"
            f"Details:\n{details}\n\n"
            f"Action items:\n"
            f"  1. Check Outlook Inbox for sales report email\n"
            f"  2. Run manually when email arrives:\n"
            f"     cd C:\\Users\\IT\\Documents\\GitHub\\tacobell-sales-dashboard\\automation\n"
            f"     .\\run_update.bat\n\n"
            f"-- KFC Sales Pipeline (auto-notification)\n"
        )
        mail.Send()
        return True
    except Exception as e:
        print(f"Failed to send notification: {e}")
        return False


CONFIG = {
    # Outlook search criteria — DAILY channel data (TACO BELL)
    "email_sender":     "Reports@foodplus.gr",
    # TB emails MUST contain "tb" or "taco" ANYWHERE in subject (case-insensitive).
    # This is more lenient than start/end matching — covers all naming variations.
    "subject_must_contain":  ["tb", "taco"],
    "email_subject":    "Weekly_Report_New",   # base subject pattern (which also includes TB somewhere)
    "attachment_hint":  "Sales Analysis",       # primary attachment hint
    "max_age_days":     7,                       # don't process emails older than this

    # Optional: HOURLY data — comes in a SEPARATE email
    "hourly_email_subject":   "Time and Day",   # partial match for hourly email subject
    "hourly_attachment_hint": "Time and Day",   # partial match for hourly attachment filename

    # Local paths — TACO BELL repo (separate from KFC)
    "repo_path":        r"C:\Users\IT\Documents\GitHub\tacobell-sales-dashboard",
    "work_dir":         r"C:\Users\IT\Documents\GitHub\tacobell-sales-dashboard\_work",
    "git_executable":   r"C:\Users\IT\AppData\Local\GitHubDesktop\app-3.4.6\resources\app\git\cmd\git.exe",
    # ↑ Αυτό το path ίσως διαφέρει στο PC σου. Αν το git δεν βρίσκεται εκεί,
    #   το script θα δοκιμάσει εναλλακτικά paths αυτόματα.

    # Budget data — αυτό αλλάζει σπάνια (ίσως 1 φορά τον χρόνο)
    "budget_source_xlsx": r"C:\Users\IT\Documents\GitHub\tacobell-sales-dashboard\_work\budget_source.xlsx",
    # Όταν αλλάξει το budget v2, απλά αντικαθιστάς αυτό το αρχείο.

    # Git authentication (PAT)
    # Το PAT διαβάζεται από τοπικό αρχείο για ασφάλεια — δεν είναι hardcoded εδώ
    "pat_file":         r"C:\Users\IT\Documents\GitHub\tacobell-sales-dashboard\_work\.github_pat",
    "git_user_name":    "Taco Bell Auto-Update Bot",
    "git_user_email":   "auto@tacobell.local",
    "github_repo":      "Reports-frt/tacobell-sales-dashboard",

    # Retry / verification settings
    "push_max_retries":      3,        # if push fails, retry this many times
    "push_retry_delay":      120,      # seconds between push retries
    "verify_max_attempts":   12,       # max polls of GitHub Pages CDN to confirm deploy
    "verify_poll_interval":  60,       # seconds between verify polls (12 × 60s = 12 minutes max wait)
    "pages_url":             "https://reports-frt.github.io/tacobell-sales-dashboard/data.json",
}

# ============================================================
# LOGGING
# ============================================================

LOG_FILE = os.path.join(CONFIG["repo_path"], "_work", "update.log")
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger(__name__)

# ============================================================
# STEP 1: GET LATEST ATTACHMENT FROM OUTLOOK
# ============================================================

def get_latest_attachment():
    """Connect to Outlook, find the most recent matching email, save attachment."""
    log.info("=" * 60)
    log.info("STEP 1: Searching Outlook for latest report...")

    try:
        import win32com.client
    except ImportError:
        log.error("pywin32 not installed. Run: pip install pywin32")
        sys.exit(1)

    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # 6 = Inbox

    sender_filter = CONFIG["email_sender"].lower()
    subject_filter = CONFIG["email_subject"].lower()
    cutoff = datetime.now() - timedelta(days=CONFIG["max_age_days"])

    # Sort by ReceivedTime descending (newest first)
    items = inbox.Items
    items.Sort("[ReceivedTime]", True)

    found_email = None
    checked = 0
    for item in items:
        checked += 1
        if checked > 200:  # don't scan whole inbox forever
            break
        try:
            received = item.ReceivedTime
            if hasattr(received, 'replace'):
                # Strip timezone for comparison (Outlook returns tz-aware)
                received_naive = received.replace(tzinfo=None)
            else:
                received_naive = received
            if received_naive < cutoff:
                break  # all older items are beyond cutoff

            sender = (item.SenderEmailAddress or "").lower()
            subject = (item.Subject or "").lower()

            # TB-only: subject MUST contain "tb" or "taco" anywhere
            required = CONFIG.get("subject_must_contain", [])
            if required and not any(s in subject for s in required):
                continue

            if sender_filter in sender and subject_filter in subject:
                if item.Attachments.Count > 0:
                    found_email = item
                    log.info(f"Found email: '{item.Subject}' from {sender} ({received_naive})")
                    break
        except Exception as e:
            log.warning(f"Error reading item: {e}")
            continue

    if not found_email:
        log.error(f"No matching email found in last {CONFIG['max_age_days']} days.")
        log.error(f"  Looking for: from='{sender_filter}' subject contains '{subject_filter}'")
        send_failure_notification(
            "Missing sales email",
            f"No email matching from='{sender_filter}' subject contains '{subject_filter}' "
            f"in the last {CONFIG['max_age_days']} days.\n"
            f"Dashboard data will remain stale until email arrives."
        )
        sys.exit(1)

    # Save the primary attachment from the channel-data email
    work_dir = CONFIG["work_dir"]
    os.makedirs(work_dir, exist_ok=True)
    saved_path = None
    primary_hint = CONFIG["attachment_hint"].lower()

    for att in found_email.Attachments:
        att_name = att.FileName.lower()
        if not att_name.endswith(('.xlsx', '.xlsm')): continue
        if primary_hint in att_name:
            saved_path = os.path.join(work_dir, "sales_analysis_latest.xlsx")
            att.SaveAsFile(saved_path)
            log.info(f"Saved primary attachment: {att.FileName} -> {saved_path}")
            break

    if not saved_path:
        log.error(f"Email found but no matching primary xlsx attachment (hint: '{primary_hint}')")
        sys.exit(1)

    # Now look for a SEPARATE email with hourly data (optional)
    saved_hourly = find_hourly_attachment(inbox, work_dir)

    return saved_path, saved_hourly


def find_hourly_attachment(folder, work_dir):
    """Search for a separate email containing hourly time-and-day data.
    Returns the saved path or None if no email found.
    
    The hourly file comes in its own email, not as a 2nd attachment in the
    weekly report email."""
    log.info("=" * 60)
    log.info("STEP 1b: Looking for separate hourly email...")

    sender = CONFIG["email_sender"].lower()
    hourly_subj_hint = CONFIG.get("hourly_email_subject", "").lower()
    hourly_att_hint = CONFIG.get("hourly_attachment_hint", "").lower()
    if not hourly_subj_hint or not hourly_att_hint:
        log.info("  (Hourly email config missing — skipping)")
        return None

    cutoff = datetime.now() - timedelta(days=CONFIG["max_age_days"])
    items = folder.Items
    items.Sort("[ReceivedTime]", True)  # newest first

    found_email = None
    checked = 0
    for it in items:
        checked += 1
        if checked > 200: break
        try:
            received = it.ReceivedTime
            if received is None: continue
            # Strip timezone for naive comparison
            if hasattr(received, "replace"):
                received = received.replace(tzinfo=None)
            if received < cutoff: break  # too old, since sorted desc

            sender_addr = ""
            try:
                sender_addr = (it.SenderEmailAddress or "").lower()
            except Exception:
                pass
            subject = (it.Subject or "").lower()

            # TB-only: subject MUST contain "tb" or "taco" anywhere
            required = CONFIG.get("subject_must_contain", [])
            if required and not any(s in subject for s in required):
                continue

            if sender in sender_addr and hourly_subj_hint in subject:
                found_email = it
                log.info(f"  Found hourly email: '{it.Subject}' from {received}")
                break
        except Exception as e:
            continue

    if not found_email:
        log.info(f"  (No hourly email found in last {CONFIG['max_age_days']} days — script will continue without hourly data)")
        return None

    saved_hourly = None
    for att in found_email.Attachments:
        att_name = att.FileName.lower()
        if not att_name.endswith(('.xlsx', '.xlsm')): continue
        if hourly_att_hint in att_name:
            saved_hourly = os.path.join(work_dir, "sales_hourly_latest.xlsx")
            att.SaveAsFile(saved_hourly)
            log.info(f"  Saved hourly attachment: {att.FileName} -> {saved_hourly}")
            break

    if not saved_hourly:
        log.warning(f"  Hourly email found but no matching xlsx attachment")
    return saved_hourly


# ============================================================
# STEP 2: PARSE EXCEL → DATA.JSON
# ============================================================

def parse_sales_analysis(xlsx_path):
    """Parse the Sales Analysis Targit export into structured records."""
    log.info("=" * 60)
    log.info("STEP 2: Parsing Sales Analysis xlsx...")

    import pandas as pd

    def prepare(df, value_col_name):
        d = df.copy()
        d.columns = ['Year', 'Month', 'Day', 'Channel', 'Store', 'Value']
        for col in ['Year', 'Month', 'Day', 'Channel']:
            d[col] = d[col].ffill()
        def to_int(x):
            try: return int(x)
            except: return None
        d['DayInt'] = d['Day'].apply(to_int)
        d['MonthInt'] = d['Month'].apply(to_int)
        d['YearInt'] = d['Year'].apply(to_int)
        mask = (d['YearInt'].notna() & d['MonthInt'].notna() & d['DayInt'].notna()
                 & (d['Channel'] != 'Συνολικά') & (d['Store'] != 'Συνολικά') & d['Store'].notna())
        leaf = d[mask].copy()
        leaf['Date'] = pd.to_datetime(dict(year=leaf['YearInt'].astype(int),
                                            month=leaf['MonthInt'].astype(int),
                                            day=leaf['DayInt'].astype(int)), errors='coerce')
        leaf = leaf[['Date', 'Store', 'Channel', 'Value']].rename(columns={'Value': value_col_name})
        return leaf

    # Find the right sheet names (might have variations)
    xl = pd.ExcelFile(xlsx_path, engine="calamine")
    log.info(f"Available sheets: {xl.sheet_names}")

    amount_sheet = next((s for s in xl.sheet_names if 'amount' in s.lower()), None)
    orders_sheet = next((s for s in xl.sheet_names if 'order' in s.lower()), None)

    if not amount_sheet or not orders_sheet:
        log.error(f"Could not find Amount/Orders sheets. Got: {xl.sheet_names}")
        sys.exit(1)

    log.info(f"  Amount sheet: '{amount_sheet}'")
    log.info(f"  Orders sheet: '{orders_sheet}'")

    df_amt = pd.read_excel(xlsx_path, sheet_name=amount_sheet, engine="calamine", header=0)
    df_ord = pd.read_excel(xlsx_path, sheet_name=orders_sheet, engine="calamine", header=0)

    amt = prepare(df_amt, 'Sales')
    ord_ = prepare(df_ord, 'TCs')
    data = amt.merge(ord_, on=['Date', 'Store', 'Channel'], how='outer')
    data['Sales'] = data['Sales'].fillna(0)
    data['TCs'] = data['TCs'].fillna(0)

    log.info(f"  Combined rows: {len(data):,}")
    log.info(f"  Date range: {data['Date'].min().date()} → {data['Date'].max().date()}")
    log.info(f"  Total Sales: {data['Sales'].sum():,.0f}€")

    return data


# ============================================================
# STEP 2b: PARSE HOURLY (optional)
# ============================================================

def parse_hourly_data(xlsx_path):
    """Parse the hourly time+day analysis file into compact aggregates.
    Returns dict with records_hsc and records_dow_hour, or None if file missing.
    
    Aggregates produced:
    - records_hsc:      [year, hour, channel_idx, store_targit_idx, sales]
                        chain-wide hour × store × channel pattern per year
    - records_dow_hour: [year, dow, hour, sales]
                        chain-wide DOW × hour pattern per year
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        log.info("STEP 2b: No hourly data file — skipping.")
        return None

    log.info("=" * 60)
    log.info("STEP 2b: Parsing hourly data...")
    import pandas as pd
    import datetime as _dt

    try:
        df = pd.read_excel(xlsx_path, engine="calamine", sheet_name=0, header=0)
    except Exception as e:
        log.warning(f"  Failed to parse hourly xlsx: {e} — skipping hourly")
        return None

    # The sheet has columns: Year Int, Month Of Year, Day Of Month, Hour..., Channels..., STORE Name, Καθαρή Αξία
    # Rename for sanity
    if len(df.columns) < 7:
        log.warning(f"  Hourly file has only {len(df.columns)} cols, expected 7 — skipping")
        return None
    df.columns = ['Year', 'Month', 'Day', 'Hour', 'Channel', 'Store', 'Sales'] + list(df.columns[7:])

    # Filter to leaf rows only (no aggregate "Συνολικά" rows)
    def _is_leaf(v):
        if v is None: return False
        s = str(v).strip()
        return s and s != 'Συνολικά'
    mask = (df['Year'].apply(_is_leaf) & df['Month'].apply(_is_leaf) &
            df['Day'].apply(_is_leaf) & df['Hour'].apply(_is_leaf) &
            df['Channel'].apply(_is_leaf) & df['Store'].apply(_is_leaf) &
            df['Sales'].notna() & (df['Channel'] != 'Unknown'))
    leaf = df[mask].copy()
    if len(leaf) == 0:
        log.warning("  No leaf rows in hourly file — skipping")
        return None

    # Type conversion
    leaf['Year'] = leaf['Year'].astype(int)
    leaf['Month'] = leaf['Month'].astype(int)
    leaf['Day'] = leaf['Day'].astype(int)
    leaf['HourInt'] = leaf['Hour'].apply(lambda x: int(str(x).split(':')[0]))

    # Merge DELIVERAS into E-FOOD GO
    leaf.loc[leaf['Channel'] == 'DELIVERAS', 'Channel'] = DELIVERAS_MERGE_INTO

    log.info(f"  Leaf rows after filtering: {len(leaf):,}")

    # Stores in this file (Targit names)
    stores_targit_in_file = sorted(leaf['Store'].unique().tolist())

    # Aggregate 1: Year × Hour × Channel × Store
    agg1 = leaf.groupby(['Year', 'HourInt', 'Channel', 'Store'], as_index=False)['Sales'].sum()

    records_hsc = []
    for _, r in agg1.iterrows():
        if r['Channel'] not in CHANNEL_ORDER:
            continue
        records_hsc.append([
            int(r['Year']),
            int(r['HourInt']),
            CHANNEL_ORDER.index(r['Channel']),
            stores_targit_in_file.index(r['Store']),
            round(float(r['Sales']), 2),
        ])

    # Aggregate 2: Year × DOW × Hour (chain-wide pattern)
    leaf['DOW'] = leaf.apply(lambda r: _dt.date(r['Year'], r['Month'], r['Day']).weekday(), axis=1)
    agg2 = leaf.groupby(['Year', 'DOW', 'HourInt'], as_index=False)['Sales'].sum()
    records_dow_hour = []
    for _, r in agg2.iterrows():
        records_dow_hour.append([int(r['Year']), int(r['DOW']), int(r['HourInt']), round(float(r['Sales']), 2)])

    # Aggregate 3: DATE × Store × Hour (full granularity for date-range filtering)
    # No channel breakdown — keeps size manageable
    leaf['DateStr'] = leaf.apply(
        lambda r: f"{int(r['Year']):04d}-{int(r['Month']):02d}-{int(r['Day']):02d}", axis=1
    )
    agg3 = leaf.groupby(['DateStr', 'Store', 'HourInt'], as_index=False)['Sales'].sum()
    
    # Build a date-index so we don't repeat ISO strings (saves bytes)
    unique_dates = sorted(agg3['DateStr'].unique().tolist())
    date_index = {d: i for i, d in enumerate(unique_dates)}
    
    records_hsd = []  # [date_idx, store_idx, hour, sales]
    for _, r in agg3.iterrows():
        records_hsd.append([
            date_index[r['DateStr']],
            stores_targit_in_file.index(r['Store']),
            int(r['HourInt']),
            round(float(r['Sales']), 2),
        ])

    log.info(f"  records_hsc (hour×store×channel×year): {len(records_hsc):,}")
    log.info(f"  records_dow_hour (DOW×hour×year):     {len(records_dow_hour):,}")
    log.info(f"  records_hsd (DATE×store×hour):        {len(records_hsd):,}")
    log.info(f"  Date range: {unique_dates[0]} → {unique_dates[-1]} ({len(unique_dates)} days)")
    log.info(f"  Years: {sorted(leaf['Year'].unique().tolist())}")
    log.info(f"  Hours: {sorted(leaf['HourInt'].unique().tolist())}")

    return {
        'years': sorted(leaf['Year'].unique().tolist()),
        'hours': sorted(leaf['HourInt'].unique().tolist()),
        'stores_targit': stores_targit_in_file,
        'records_hsc': records_hsc,
        'records_dow_hour': records_dow_hour,
        'dates': unique_dates,
        'records_hsd': records_hsd,
    }


# ============================================================
# STEP 3: BUILD JSON
# ============================================================

# Static mappings (these don't change) — derived from Help sheet of Management Report
# Each store: list of Targit name variants, plus metadata.
# When matching channel records to stores, comparison is case-insensitive AND ignores extra whitespace.
# Just add new variants to TargitNames list when needed.
#
# StoreType values (from Help sheet):
#   "In Line"          - free-standing inline location
#   "In Line Mall"     - in mall, inline (separate entrance)
#   "Mall Food Court"  - inside food court
#   "Drive Through"    - dedicated drive-through location
STORE_MAPPING = [
    # TB stores will be auto-registered from sales data on first encounter.
    # To override defaults (store_type, opening_date, is_new), add explicit entries here:
    # ([TargitNames], ResultName, ShortLabel, StoreType, OpeningDate, IsNewStore)
]
# When new stores open or their classification changes (Same -> mature etc), edit the IsNewStore flag here.

STORE_TYPES = ["In Line", "In Line Mall", "Mall Food Court", "Drive Through"]

# DELIVERAS is essentially zero - merge it into E-FOOD GO at parse time.
DELIVERAS_MERGE_INTO = 'E-FOOD GO'

# All distinct individual channels (after DELIVERAS merge into E-FOOD GO)
CHANNEL_ORDER = ['KIOSK', 'STORE', 'MOBILE', 'WEB SITE', 'DRIVETHRU', 'E-FOOD GO', 'WOLT', 'BOX']

# View A: "Type" — operational view (3 groups)
# DELIVERAS is essentially gone (€0.44 in dataset). Merge into E-FOOD GO at parse time.
DELIVERAS_MERGE_INTO = 'E-FOOD GO'

TYPE_ORDER = ['STORE', 'DRIVE THRU', 'END-TO-END']

# View A — "Type" (3-way operational split)
#   STORE      = anything done at the store (KIOSK + STORE cashier + MOBILE pickup + WEB SITE pickup)
#   DRIVE THRU = drive-through window
#   END-TO-END = delivery aggregators (E-FOOD GO + WOLT + BOX)
# NOTE: MOBILE & WEB SITE are app/web orders where customer comes to store for pickup.
#       They are NOT delivery — they are STORE.
CHANNEL_TO_TYPE = {
    'KIOSK': 'STORE', 'STORE': 'STORE',
    'MOBILE': 'STORE', 'WEB SITE': 'STORE',
    'DRIVETHRU': 'DRIVE THRU',
    'E-FOOD GO': 'END-TO-END', 'WOLT': 'END-TO-END', 'BOX': 'END-TO-END',
    'Unknown': 'STORE',
}

# View B — "Sales Type" (4-way customer-experience split)
#   DINEIN     = customer eats inside (KIOSK self-service)
#   TAKEAWAY   = ordered at the store but takes away (cashier STORE + MOBILE + WEB SITE)
#   DRIVE THRU = ordered through drive-through window
#   END-TO-END = ordered through delivery aggregators
CHANNEL_TO_SALESTYPE = {
    'KIOSK': 'DINEIN',
    'STORE': 'TAKEAWAY', 'MOBILE': 'TAKEAWAY', 'WEB SITE': 'TAKEAWAY',
    'DRIVETHRU': 'DRIVE THRU',
    'E-FOOD GO': 'END-TO-END', 'WOLT': 'END-TO-END', 'BOX': 'END-TO-END',
    'Unknown': 'TAKEAWAY',
}

# Order in which channels appear in dashboard tables
CHANNEL_ORDER = ['KIOSK', 'STORE', 'MOBILE', 'WEB SITE', 'DRIVETHRU', 'E-FOOD GO', 'WOLT', 'BOX']
TYPE_ORDER = ['STORE', 'DRIVE THRU', 'END-TO-END']
SALESTYPE_ORDER = ['DINEIN', 'TAKEAWAY', 'DRIVE THRU', 'END-TO-END']  # for View B


def load_budget(budget_xlsx_path):
    """Read Budget v1 from the Management Report 'Result excl. All EORD' sheet."""
    log.info("=" * 60)
    log.info(f"STEP 3a: Loading Budget from {budget_xlsx_path}")

    if not os.path.exists(budget_xlsx_path):
        log.error(f"Budget source not found: {budget_xlsx_path}")
        log.error("Place the Management Report xlsx at this path and re-run.")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(budget_xlsx_path, data_only=True)

    # Find 'Result excl. All EORD' sheet (or fallback)
    result_sheet_name = None
    for s in wb.sheetnames:
        if 'result' in s.lower() and 'eord' in s.lower():
            result_sheet_name = s
            break
    if not result_sheet_name:
        result_sheet_name = next((s for s in wb.sheetnames if 'result' in s.lower()), None)
    if not result_sheet_name:
        log.error(f"Could not find Result sheet in budget file. Sheets: {wb.sheetnames}")
        sys.exit(1)

    log.info(f"  Reading sheet: '{result_sheet_name}'")
    ws = wb[result_sheet_name]

    # Auto-detect store column: try col 2 first (new format), fallback col 1 (old format).
    # Detect by counting how many distinct values look like "KFC ..." in each column.
    def count_kfc_values(col_idx):
        seen = set()
        for r in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=r, column=col_idx).value
            if v and str(v).strip().upper().startswith('KFC'):
                seen.add(str(v).strip())
        return len(seen)
    c1, c2 = count_kfc_values(1), count_kfc_values(2)
    store_col = 2 if c2 > c1 else 1
    log.info(f"  Store name column: {store_col} (col 1 has {c1} KFC names, col 2 has {c2})")

    # Auto-detect month columns by scanning row 4 (dates).
    # Row 4 contains datetime objects for each monthly column.
    import datetime as _dt
    month_cols = {}  # month_int (1-12) -> column index
    for c in range(6, 25):
        v = ws.cell(row=4, column=c).value
        if isinstance(v, _dt.datetime):
            month_cols[v.month] = c
    if len(month_cols) < 12:
        log.warning(f"  Only found {len(month_cols)} month columns in row 4. Months found: {sorted(month_cols.keys())}")
    else:
        log.info(f"  Month columns detected: {month_cols}")

    budget = {}
    for r in range(5, ws.max_row + 1):
        store = ws.cell(row=r, column=store_col).value
        desc = ws.cell(row=r, column=5).value
        if not store or not desc:
            continue
        store = str(store).strip()
        desc = str(desc).strip()
        if store not in budget:
            budget[store] = {m: {} for m in range(1, 13)}
        if desc == 'Sales, Total':
            for m, c in month_cols.items():
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    budget[store][m]['Sales'] = float(v) * 1000  # in thousands
        elif desc == 'Transactions, Total':
            for m, c in month_cols.items():
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    budget[store][m]['TCs'] = float(v) * 1000
        elif desc == 'Ave Spend, Total':
            for m, c in month_cols.items():
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    budget[store][m]['AvgTicket'] = float(v)

    log.info(f"  Budget loaded for {len(budget)} stores")
    return budget


def build_data_json(data, budget, hourly_data=None):
    """Combine sales data + budget into the dashboard JSON format.
    
    Emits TWO channel views:
      - records_type: 3-way operational split (STORE / DRIVE THRU / END-TO-END)
      - records_st:   4-way customer split   (DINEIN / TAKEAWAY / DRIVE THRU / END-TO-END)
    Plus per-channel breakdown:
      - records_ch:   8 individual channels (KIOSK, STORE, MOBILE, WEB SITE,
                                              DRIVETHRU, E-FOOD GO, WOLT, BOX)
                      DELIVERAS rows are merged into E-FOOD GO before aggregation.
    """

    # ── PATCH: Amount/Sales column compatibility ──────────────────────
    # parse_sales_analysis produces 'Sales'; legacy code uses 'Amount'.
    # Ensure both column names exist as aliases so all downstream code works.
    if 'Sales' in data.columns and 'Amount' not in data.columns:
        data = data.copy()
        data['Amount'] = data['Sales']
    elif 'Amount' in data.columns and 'Sales' not in data.columns:
        data = data.copy()
        data['Sales'] = data['Amount']
    # ──────────────────────────────────────────────────────────────────

    log.info("=" * 60)
    log.info("STEP 3b: Building data.json structure...")

    # Normalize for matching
    def _norm(s):
        if s is None: return ''
        return ' '.join(str(s).upper().split())

    # Build mapping tables from new 6-tuple STORE_MAPPING
    targit_to_result = {}
    for entry in STORE_MAPPING:
        targit_names, result_name = entry[0], entry[1]
        for tn in targit_names:
            targit_to_result[_norm(tn)] = result_name

    result_to_meta = {}
    store_order = []
    for entry in STORE_MAPPING:
        targit_names, result_name, short_label, store_type, opening, is_new = entry
        result_to_meta[result_name] = {
            'short': short_label,
            'targit': targit_names[0],
            'store_type': store_type,
            'opening_date': opening,
            'is_new': bool(is_new),
        }
        store_order.append(result_name)

    # Map Targit names → Result names
    data['StoreResult'] = data['Store'].apply(lambda s: targit_to_result.get(_norm(s)))
    
    # Auto-register unknown stores (so we don't lose data when new stores open).
    # Uses default metadata (In Line, today's date, marked as new).
    unmapped = data[data['StoreResult'].isna()]['Store'].unique()
    if len(unmapped) > 0:
        log.info(f"Auto-registering {len(unmapped)} new stores: {list(unmapped)}")
        from datetime import date as _date
        today_str = _date.today().isoformat()
        for raw_name in unmapped:
            # Clean store name for use as both ResultName and ShortLabel
            # E.g., "TACO BELL- ΧΑΛΑΝΔΡΙ" → result "TB HALANDRI", short "ΧΑΛΑΝΔΡΙ"
            cleaned = str(raw_name).strip()
            # Strip brand prefix (KFC -, TACO BELL-, TB -, etc) for short label
            short = cleaned
            # Include Greek Tau (Τ) variants — Targit sometimes uses Greek Τ instead of Latin T
            for prefix in [
                'TACO BELL- ', 'TACO BELL-', 'TACO BELL ',
                'ΤACO BELL- ', 'ΤACO BELL-', 'ΤACO BELL ',  # Greek Tau
                'KFC - ', 'KFC- ', 'KFC-', 'KFC ',
                'TB - ', 'TB- ', 'TB-', 'TB ',
            ]:
                if short.upper().startswith(prefix.upper()):
                    short = short[len(prefix):].strip()
                    break
            # Use the original (cleaned) name as the Result name — preserves brand identity
            result_name = cleaned
            result_to_meta[result_name] = {
                'short': short or cleaned,
                'targit': cleaned,
                'store_type': 'In Line',  # default — manually edit STORE_MAPPING for precise type
                'opening_date': today_str,
                'is_new': True,
            }
            store_order.append(result_name)
            targit_to_result[_norm(cleaned)] = result_name
            log.info(f"  + {cleaned}  →  short='{short}', opened='{today_str}' (default metadata; edit STORE_MAPPING for precise values)")
        # Re-apply the mapping with the new entries
        data['StoreResult'] = data['Store'].apply(lambda s: targit_to_result.get(_norm(s)))
    data = data[data['StoreResult'].notna()].copy()

    # Merge DELIVERAS into E-FOOD GO at parse time (DELIVERAS is essentially defunct).
    deliveras_count = (data['Channel'] == 'DELIVERAS').sum()
    if deliveras_count > 0:
        log.info(f"  Merging {deliveras_count} DELIVERAS rows into {DELIVERAS_MERGE_INTO}")
        data.loc[data['Channel'] == 'DELIVERAS', 'Channel'] = DELIVERAS_MERGE_INTO

    # Map each row to its Sales Type (View B), Type (View A), and keep raw Channel
    data['SalesType'] = data['Channel'].map(CHANNEL_TO_SALESTYPE)
    data['Type'] = data['Channel'].map(CHANNEL_TO_TYPE)

    # ---- Daily aggregates at THREE granularities ----
    daily_st   = data.groupby(['Date', 'StoreResult', 'SalesType'], as_index=False)[['Sales', 'TCs']].sum()
    daily_type = data.groupby(['Date', 'StoreResult', 'Type'],      as_index=False)[['Sales', 'TCs']].sum()
    daily_ch   = data.groupby(['Date', 'StoreResult', 'Channel'],   as_index=False)[['Sales', 'TCs']].sum()

    store_idx = {s: i for i, s in enumerate(store_order)}
    st_idx   = {s: i for i, s in enumerate(SALESTYPE_ORDER)}
    type_idx = {s: i for i, s in enumerate(TYPE_ORDER)}
    ch_idx   = {c: i for i, c in enumerate(CHANNEL_ORDER)}

    def to_records(df, col_idx_map, col_name):
        out = []
        for _, row in df.iterrows():
            if row[col_name] not in col_idx_map:
                continue
            out.append([
                row['Date'].strftime('%Y-%m-%d'),
                store_idx[row['StoreResult']],
                col_idx_map[row[col_name]],
                round(float(row['Sales']), 2),
                int(row['TCs']),
            ])
        return out

    records_st   = to_records(daily_st,   st_idx,   'SalesType')
    records_type = to_records(daily_type, type_idx, 'Type')
    records_ch   = to_records(daily_ch,   ch_idx,   'Channel')

    log.info(f"  records_st (Sales Type, View B):  {len(records_st):,}")
    log.info(f"  records_type (Type, View A):      {len(records_type):,}")
    log.info(f"  records_ch (individual channels): {len(records_ch):,}")

    # Budget structured
    budget_out = {}
    for store in store_order + ['KFC ALL']:
        if store in budget:
            budget_out[store] = {}
            for m in range(1, 13):
                mb = budget[store].get(m, {})
                if mb:
                    budget_out[store][str(m)] = {
                        'Sales': round(mb.get('Sales', 0), 2),
                        'TCs': round(mb.get('TCs', 0)),
                        'AvgTicket': round(mb.get('AvgTicket', 0), 2),
                    }

    # Store meta — full info for dashboard filtering
    stores_meta = dict(result_to_meta)
    stores_meta['KFC ALL'] = {
        'short': 'KFC ALL', 'targit': 'KFC ALL',
        'store_type': 'All', 'opening_date': '', 'is_new': False,
    }

    dates = sorted({r[0] for r in records_st})
    output = {
        'meta': {
            'latest_date':  max(dates) if dates else None,
            'first_date':   min(dates) if dates else None,
            'stores':       store_order,
            'salestypes':   SALESTYPE_ORDER,
            'types':        TYPE_ORDER,
            'channels':     CHANNEL_ORDER,
            'store_meta':   stores_meta,
            'store_types':  list(sorted({m['store_type'] for m in result_to_meta.values()})),
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        },
        'budget':       budget_out,
        # Year-keyed budget archive (preserved across years by merge_with_existing)
        'budget_by_year': { str(datetime.now().year): budget_out },
        'records_st':   records_st,
        'records_type': records_type,
        'records_ch':   records_ch,
    }

    # If hourly data was provided, embed it (re-mapped to result store names)
    if hourly_data:
        hourly_store_idx = []
        unmapped_h = []
        for tgt in hourly_data['stores_targit']:
            res = targit_to_result.get(_norm(tgt))
            if res and res in store_idx:
                hourly_store_idx.append(store_idx[res])
            else:
                hourly_store_idx.append(-1)
                unmapped_h.append(tgt)
        if unmapped_h:
            log.warning(f"  Hourly data has unmapped stores: {unmapped_h}")
        # Remap each record's store index from "hourly file index" → "main store index"
        records_hsc_remapped = []
        for r in hourly_data['records_hsc']:
            si_main = hourly_store_idx[r[3]]
            if si_main >= 0:
                records_hsc_remapped.append([r[0], r[1], r[2], si_main, r[4]])
        # Also remap records_hsd (date × store × hour)
        records_hsd_remapped = []
        if 'records_hsd' in hourly_data:
            for r in hourly_data['records_hsd']:
                # r = [date_idx, store_idx, hour, sales]
                si_main = hourly_store_idx[r[1]]
                if si_main >= 0:
                    records_hsd_remapped.append([r[0], si_main, r[2], r[3]])
        output['hourly'] = {
            'years':            hourly_data['years'],
            'hours':            hourly_data['hours'],
            'records_hsc':      records_hsc_remapped,
            'records_dow_hour': hourly_data['records_dow_hour'],
            'dates':            hourly_data.get('dates', []),
            'records_hsd':      records_hsd_remapped,
        }
        log.info(f"  Embedded hourly: {len(records_hsc_remapped)} hsc + {len(records_hsd_remapped)} hsd records")

    log.info(f"  Latest date: {output['meta']['latest_date']}")
    log.info(f"  Stores: {len(store_order)}")
    log.info(f"  Channels: {len(CHANNEL_ORDER)}, Types: {len(TYPE_ORDER)}, SalesTypes: {len(SALESTYPE_ORDER)}")
    return output


# ============================================================
# STEP 4: GIT PUSH
# ============================================================

def find_git_executable():
    """Locate the git binary, trying multiple known paths."""
    # First, try the configured path
    configured = CONFIG.get("git_executable")
    if configured and os.path.exists(configured):
        return configured

    # Common GitHub Desktop install locations
    user_profile = os.environ.get('USERPROFILE', '')
    candidates = []
    github_local = os.path.join(user_profile, 'AppData', 'Local', 'GitHubDesktop')
    if os.path.exists(github_local):
        # Find any app-* folder
        for entry in os.listdir(github_local):
            if entry.startswith('app-'):
                candidate = os.path.join(github_local, entry, 'resources', 'app', 'git', 'cmd', 'git.exe')
                if os.path.exists(candidate):
                    candidates.append(candidate)

    # Standard Git for Windows locations
    candidates.extend([
        r"C:\Program Files\Git\cmd\git.exe",
        r"C:\Program Files (x86)\Git\cmd\git.exe",
    ])

    for c in candidates:
        if os.path.exists(c):
            return c

    # Last resort — assume it's in PATH
    return "git"


def preserve_previous_hourly(json_data, json_path):
    """If the new json_data has no 'hourly' block but the existing data.json
    on disk DOES have one, preserve the existing block to avoid losing
    hourly data when the hourly email is delayed/missing for a day.
    
    This is mutating json_data in place.
    """
    if 'hourly' in json_data:
        return  # we have fresh hourly data, no need to preserve

    if not os.path.exists(json_path):
        return  # no previous file to read from

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            prev = json.load(f)
    except Exception as e:
        log.warning(f"  Could not read previous data.json for hourly preservation: {e}")
        return

    if 'hourly' not in prev:
        return  # previous file had no hourly either

    json_data['hourly'] = prev['hourly']
    n_hsc = len(prev['hourly'].get('records_hsc', []))
    n_dh = len(prev['hourly'].get('records_dow_hour', []))
    n_hsd = len(prev['hourly'].get('records_hsd', []))
    yrs = prev['hourly'].get('years', [])
    log.info(f"  Preserved previous hourly block: {n_hsc} hsc + {n_dh} dow-hour + {n_hsd} hsd records, years {yrs}")


def git_push_data_json(json_data):
    """Write data.json to the repo and push to GitHub."""
    log.info("=" * 60)
    log.info("STEP 4: Pushing to GitHub...")

    repo = CONFIG["repo_path"]
    json_path = os.path.join(repo, "data.json")

    # If hourly data is missing from this run but existed in the previous data.json,
    # carry it forward so the dashboard's Hourly tab doesn't disappear.
    preserve_previous_hourly(json_data, json_path)

    # ===== Step A: Backup current data.json =====
    backup_path = None
    if os.path.exists(json_path):
        from datetime import datetime as _dt
        ts = _dt.now().strftime('%Y%m%d-%H%M%S')
        backup_path = os.path.join(repo, f'data.json.backup-{ts}')
        try:
            import shutil
            shutil.copy(json_path, backup_path)
            log.info(f"  ✓ Backed up to {os.path.basename(backup_path)}")
        except Exception as e:
            log.warning(f"  Backup failed: {e}")
            backup_path = None

    # ===== Step B: Merge with existing data.json (preserves historical) =====
    json_data = merge_with_existing(json_data, json_path)

    # ===== Step C: Pre-write validation =====
    ok, issues = validate_data_json_structure(json_data)
    if not ok:
        log.error("  ❌ Pre-write validation FAILED:")
        for issue in issues:
            log.error(f"     - {issue}")
        log.error("  ABORTING write to protect data.json from corruption.")
        sys.exit(1)

    # ===== Step D: Atomic write (write-temp, rename) =====
    tmp_path = json_path + '.tmp'
    try:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, separators=(',', ':'), ensure_ascii=False)
        os.replace(tmp_path, json_path)
        size_mb = os.path.getsize(json_path) / 1024 / 1024
        log.info(f"  ✓ Wrote data.json ({size_mb:.2f} MB)")
    except Exception as e:
        log.error(f"  ❌ Write failed: {e}")
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except: pass
        if backup_path and os.path.exists(backup_path):
            log.info(f"  Restoring backup {os.path.basename(backup_path)}...")
            import shutil
            shutil.copy(backup_path, json_path)
        sys.exit(1)

    # ===== Step E: Post-write validation =====
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            re_read = json.load(f)
        ok, issues = validate_data_json_structure(re_read)
        if not ok:
            log.error("  ❌ Post-write validation FAILED:")
            for issue in issues:
                log.error(f"     - {issue}")
            if backup_path and os.path.exists(backup_path):
                log.error(f"  Rolling back to {os.path.basename(backup_path)}...")
                import shutil
                shutil.copy(backup_path, json_path)
            sys.exit(1)
        log.info("  ✓ Post-write validation passed")
    except Exception as e:
        log.error(f"  Could not re-read data.json: {e}")
        if backup_path and os.path.exists(backup_path):
            import shutil
            shutil.copy(backup_path, json_path)
        sys.exit(1)
    
    # ===== Step F: Cleanup old backups (keep last 10) =====
    cleanup_old_backups(repo, keep_n=10)

    git_exe = find_git_executable()
    log.info(f"  Using git: {git_exe}")

    # Read PAT
    pat_file = CONFIG["pat_file"]
    if not os.path.exists(pat_file):
        log.error(f"PAT file not found: {pat_file}")
        log.error("Create this file with the GitHub Personal Access Token as its only content.")
        sys.exit(1)
    with open(pat_file, 'r', encoding='utf-8') as f:
        pat = f.read().strip()

    repo_full = CONFIG["github_repo"]
    push_url = f"https://x-access-token:{pat}@github.com/{repo_full}.git"

    def run(cmd, allow_fail=False):
        log.info(f"  $ git {' '.join(cmd[1:])}")
        result = subprocess.run(
            cmd, cwd=repo, capture_output=True, text=True,
            encoding='utf-8', errors='replace'
        )
        if result.stdout.strip():
            log.info(f"    {result.stdout.strip()}")
        if result.stderr.strip():
            # git logs many things to stderr; only flag as error if it failed
            if result.returncode != 0:
                log.error(f"    {result.stderr.strip()}")
            else:
                log.info(f"    {result.stderr.strip()}")
        if result.returncode != 0 and not allow_fail:
            log.error(f"git command failed (exit {result.returncode})")
            sys.exit(1)
        return result

    # Configure user identity (idempotent)
    run([git_exe, "config", "user.name", CONFIG["git_user_name"]])
    run([git_exe, "config", "user.email", CONFIG["git_user_email"]])

    # Stage and check if anything changed
    run([git_exe, "add", "data.json"])
    diff = subprocess.run(
        [git_exe, "diff", "--cached", "--quiet"],
        cwd=repo
    )
    if diff.returncode == 0:
        log.info("  No changes in data.json — skipping push.")
        return

    # Commit
    today = datetime.now().strftime('%Y-%m-%d %H:%M')
    latest = json_data['meta']['latest_date']
    commit_msg = f"Auto-update: data through {latest} (run {today})"
    run([git_exe, "commit", "-m", commit_msg])

    # Push using PAT-injected URL with retry on transient failure
    max_retries = CONFIG.get("push_max_retries", 3)
    retry_delay = CONFIG.get("push_retry_delay", 120)
    last_err = None
    for attempt in range(1, max_retries + 1):
        if attempt > 1:
            log.info(f"  Retry attempt {attempt}/{max_retries} after {retry_delay}s wait...")
            import time
            time.sleep(retry_delay)
        push_result = subprocess.run(
            [git_exe, "push", push_url, "HEAD:main"],
            cwd=repo, capture_output=True, text=True,
            encoding='utf-8', errors='replace', timeout=120
        )
        # Mask the token in any logged output
        masked_stdout = (push_result.stdout or "").replace(pat, "***TOKEN***")
        masked_stderr = (push_result.stderr or "").replace(pat, "***TOKEN***")
        if masked_stdout.strip():
            log.info(f"    {masked_stdout.strip()}")
        if masked_stderr.strip():
            if push_result.returncode != 0:
                log.warning(f"    {masked_stderr.strip()}")
            else:
                log.info(f"    {masked_stderr.strip()}")
        if push_result.returncode == 0:
            log.info("  ✓ Pushed to GitHub successfully.")
            return  # SUCCESS — exit function
        last_err = masked_stderr or masked_stdout
        log.warning(f"  Push attempt {attempt} failed (exit {push_result.returncode}). {('Will retry...' if attempt < max_retries else 'Out of retries.')}")
    log.error(f"All {max_retries} push attempts failed. Last error: {last_err[:200] if last_err else 'unknown'}")
    sys.exit(1)


def verify_deployment(json_data):
    """Poll the GitHub Pages CDN to confirm the new data.json is live.
    
    GitHub Pages can take 1-10 minutes to propagate after a push,
    especially when GitHub Actions is having issues. We poll the deployed
    URL until we see the new latest_date, with timeout."""
    log.info("=" * 60)
    log.info("STEP 5: Verifying GitHub Pages deployment...")
    
    pages_url = CONFIG.get("pages_url")
    if not pages_url:
        log.warning("  pages_url not configured — skipping verification")
        return
    
    expected_latest = json_data['meta']['latest_date']
    expected_generated = json_data['meta']['generated_at']
    log.info(f"  Expected latest_date: {expected_latest}")
    log.info(f"  Expected generated_at: {expected_generated}")
    log.info(f"  Polling: {pages_url}")
    
    max_attempts = CONFIG.get("verify_max_attempts", 12)
    poll_interval = CONFIG.get("verify_poll_interval", 60)
    
    import urllib.request
    import time
    
    for attempt in range(1, max_attempts + 1):
        try:
            # Add cache-buster to bypass any CDN caching
            url_with_buster = pages_url + "?v=" + str(int(time.time()))
            req = urllib.request.Request(url_with_buster, headers={
                'Cache-Control': 'no-cache',
                'Pragma': 'no-cache',
                'User-Agent': 'KFC-Dashboard-Verifier/1.0'
            })
            with urllib.request.urlopen(req, timeout=20) as resp:
                # Parse just the meta block — don't load full 2MB JSON repeatedly
                # Read first 8KB which contains the meta object
                first_chunk = resp.read(8192).decode('utf-8', errors='replace')
                # Look for our markers
                if expected_generated in first_chunk:
                    log.info(f"  ✓ Deployment verified on attempt {attempt} (generated_at matches)")
                    return
                elif expected_latest in first_chunk and '"latest_date":"' + expected_latest + '"' in first_chunk:
                    log.info(f"  ✓ Deployment verified on attempt {attempt} (latest_date matches)")
                    return
                else:
                    # Try to extract the live latest_date for diagnostic
                    import re
                    m = re.search(r'"latest_date":"([^"]+)"', first_chunk)
                    live_date = m.group(1) if m else "?"
                    log.info(f"  Attempt {attempt}/{max_attempts}: not yet deployed (live latest_date: {live_date})")
        except Exception as e:
            log.info(f"  Attempt {attempt}/{max_attempts}: fetch failed ({type(e).__name__}: {str(e)[:100]})")
        
        if attempt < max_attempts:
            time.sleep(poll_interval)
    
    log.warning(f"  ⚠ Deployment not verified after {max_attempts} attempts ({max_attempts*poll_interval}s).")
    log.warning(f"     The push succeeded but GitHub Pages may still be deploying.")
    log.warning(f"     Check https://github.com/{CONFIG['github_repo']}/actions for status.")
    # Don't exit with error — push was OK, deployment is just slow.


# ============================================================
# MAIN
# ============================================================

def merge_with_existing(new_json, json_path):
    """Merge newly-built JSON with existing data.json to preserve historical data.
    
    CRITICAL for bullet-proofing: pipeline used to overwrite data.json every run,
    losing historical data. This function:
      1. Reads existing data.json (if present)
      2. Merges historical records that the new build doesn't include
      3. Returns unified JSON with FULL history preserved
    
    Dedup key: (date, store_idx, sub_idx). New data wins on conflicts.
    """
    if not os.path.exists(json_path):
        log.info("  No existing data.json — nothing to merge")
        return new_json
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            existing = json.load(f)
    except Exception as e:
        log.warning(f"  Could not read existing data.json: {e}")
        log.warning("  Proceeding with new data only (no merge)")
        return new_json
    
    log.info("=" * 60)
    log.info("STEP 3c: Merging with existing data.json (preserve historical)")
    
    existing_st = existing.get('records_st', [])
    new_st = new_json.get('records_st', [])
    
    if not existing_st:
        log.info("  Existing data.json has no records — no merge needed")
        return new_json
    
    existing_dates = set(r[0] for r in existing_st)
    new_dates = set(r[0] for r in new_st)
    new_min = min(new_dates) if new_dates else None
    new_max = max(new_dates) if new_dates else None
    existing_min = min(existing_dates) if existing_dates else None
    existing_max = max(existing_dates) if existing_dates else None
    
    log.info(f"  Existing data: {len(existing_st):,} records, {existing_min} → {existing_max}")
    log.info(f"  New data:      {len(new_st):,} records, {new_min} → {new_max}")
    
    e_meta = existing.get('meta', {})
    n_meta = new_json.get('meta', {})
    e_stores = e_meta.get('stores', [])
    n_stores = n_meta.get('stores', [])
    
    only_in_existing = [s for s in e_stores if s not in n_stores]
    only_in_new = [s for s in n_stores if s not in e_stores]
    
    if only_in_existing:
        log.warning(f"  Stores in existing not in new: {only_in_existing} (preserving)")
    if only_in_new:
        log.info(f"  New stores in new data: {only_in_new}")
    
    unified_stores = list(e_stores) + [s for s in n_stores if s not in e_stores]
    e_si_to_unified = {i: unified_stores.index(s) for i, s in enumerate(e_stores)}
    n_si_to_unified = {i: unified_stores.index(s) for i, s in enumerate(n_stores)}
    
    for key in ['salestypes', 'types', 'channels']:
        e_v = e_meta.get(key, [])
        n_v = n_meta.get(key, [])
        if e_v and n_v and e_v != n_v:
            log.error(f"  Schema mismatch on '{key}'!")
            log.error(f"    Existing: {e_v}")
            log.error(f"    New:      {n_v}")
            log.error(f"  Aborting merge to avoid data corruption. Using new data only.")
            return new_json
    
    for record_key in ['records_st', 'records_type', 'records_ch']:
        e_records = existing.get(record_key, [])
        n_records = new_json.get(record_key, [])
        
        remapped_existing = []
        for r in e_records:
            if len(r) < 5:
                continue
            d, si, sub, s, t = r
            new_si = e_si_to_unified.get(si)
            if new_si is None:
                continue
            remapped_existing.append([d, new_si, sub, s, t])
        
        remapped_new = []
        for r in n_records:
            if len(r) < 5:
                continue
            d, si, sub, s, t = r
            new_si = n_si_to_unified.get(si)
            if new_si is None:
                continue
            remapped_new.append([d, new_si, sub, s, t])
        
        new_keys = set((r[0], r[1], r[2]) for r in remapped_new)
        kept_existing = [r for r in remapped_existing if (r[0], r[1], r[2]) not in new_keys]
        
        merged = kept_existing + remapped_new
        merged.sort(key=lambda r: (r[0], r[1], r[2]))
        
        new_json[record_key] = merged
        log.info(f"  {record_key}: +{len(remapped_new):,} new | preserved {len(kept_existing):,} historical | total {len(merged):,}")
    
    new_json['meta']['stores'] = unified_stores
    
    e_store_meta = e_meta.get('store_meta', {})
    n_store_meta = new_json.get('meta', {}).get('store_meta', {})
    for sname, smeta in e_store_meta.items():
        if sname not in n_store_meta:
            n_store_meta[sname] = smeta
    new_json['meta']['store_meta'] = n_store_meta
    
    # Merge hourly
    e_hourly = existing.get('hourly')
    n_hourly = new_json.get('hourly')
    
    if e_hourly and isinstance(e_hourly, dict):
        if not n_hourly or not isinstance(n_hourly, dict):
            log.info("  New data has no hourly — preserving existing hourly")
            new_json['hourly'] = e_hourly
        else:
            log.info("  Merging hourly data (year-level dedup)...")
            for hkey in ['records_hsc', 'records_dow_hour']:
                e_arr = e_hourly.get(hkey, [])
                n_arr = n_hourly.get(hkey, [])
                n_years = set(r[0] for r in n_arr if r)
                preserved = [r for r in e_arr if r and r[0] not in n_years]
                n_hourly[hkey] = preserved + n_arr
                log.info(f"    {hkey}: +{len(n_arr):,} new | preserved {len(preserved):,} historical | total {len(n_hourly[hkey]):,}")
            
            e_dates = list(e_hourly.get('dates', []))
            n_dates = list(n_hourly.get('dates', []))
            unified_dates = sorted(set(e_dates) | set(n_dates))
            date_to_unified_idx = {d: i for i, d in enumerate(unified_dates)}
            
            rebuilt_hsd = []
            new_hsd_keys = set()
            
            for r in n_hourly.get('records_hsd', []):
                if not r or len(r) < 4:
                    continue
                old_di, si, hr, sl = r[0], r[1], r[2], r[3]
                if old_di >= len(n_dates):
                    continue
                d = n_dates[old_di]
                new_di = date_to_unified_idx.get(d)
                new_si = n_si_to_unified.get(si)
                if new_di is None or new_si is None:
                    continue
                new_hsd_keys.add((new_di, new_si, hr))
                rebuilt_hsd.append([new_di, new_si, hr, sl])
            
            for r in e_hourly.get('records_hsd', []):
                if not r or len(r) < 4:
                    continue
                old_di, si, hr, sl = r[0], r[1], r[2], r[3]
                if old_di >= len(e_dates):
                    continue
                d = e_dates[old_di]
                new_di = date_to_unified_idx.get(d)
                new_si = e_si_to_unified.get(si)
                if new_di is None or new_si is None:
                    continue
                if (new_di, new_si, hr) in new_hsd_keys:
                    continue
                rebuilt_hsd.append([new_di, new_si, hr, sl])
            
            n_hourly['dates'] = unified_dates
            n_hourly['records_hsd'] = rebuilt_hsd
            n_hourly['years'] = sorted(set(n_hourly.get('years', [])) | set(e_hourly.get('years', [])))
            n_hourly['hours'] = sorted(set(n_hourly.get('hours', [])) | set(e_hourly.get('hours', [])))
            log.info(f"    records_hsd: total {len(rebuilt_hsd):,} across {len(unified_dates):,} dates")
            log.info(f"    Final years: {n_hourly['years']}")
            
            new_json['hourly'] = n_hourly
    
    # ===== Preserve budget archive (budget_by_year) =====
    e_bby = existing.get('budget_by_year', {}) or {}
    n_bby = new_json.get('budget_by_year', {}) or {}
    merged_bby = dict(e_bby)
    merged_bby.update(n_bby)
    if merged_bby:
        new_json['budget_by_year'] = merged_bby
        log.info(f"  budget_by_year: years preserved = {sorted(merged_bby.keys())}")

    all_dates = [r[0] for r in new_json.get('records_st', [])]
    if all_dates:
        new_json['meta']['first_date'] = min(all_dates)
        new_json['meta']['latest_date'] = max(all_dates)
    
    log.info(f"  ✓ Merge complete: {new_json['meta'].get('first_date')} → {new_json['meta'].get('latest_date')}")
    return new_json


def validate_data_json_structure(data):
    """Sanity-check data.json before/after writing. Returns (ok, issues)."""
    issues = []
    for k in ['meta', 'records_st', 'records_ch']:
        if k not in data:
            issues.append(f"Missing top-level key: {k}")
    if 'meta' in data:
        meta = data['meta']
        for k in ['stores', 'salestypes', 'channels']:
            if k not in meta:
                issues.append(f"Missing meta key: {k}")
    for k in ['records_st', 'records_type', 'records_ch']:
        if k in data:
            records = data[k]
            if not records:
                continue
            sample = records[0]
            if len(sample) != 5:
                issues.append(f"{k}: row length is {len(sample)}, expected 5. Sample: {sample}")
            n_stores = len(data.get('meta', {}).get('stores', []))
            if n_stores > 0:
                max_si = max(r[1] for r in records if len(r) >= 2)
                if max_si >= n_stores:
                    issues.append(f"{k}: max store_idx {max_si} >= store count {n_stores}")
    return len(issues) == 0, issues


def cleanup_old_backups(repo_path, keep_n=10):
    """Keep only the N most recent data.json.backup-* files in repo root."""
    import glob
    pattern = os.path.join(repo_path, 'data.json.backup-*')
    backups = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if len(backups) > keep_n:
        for old in backups[keep_n:]:
            try:
                os.remove(old)
                log.info(f"  Cleaned up old backup: {os.path.basename(old)}")
            except Exception as e:
                log.warning(f"  Failed to remove backup {old}: {e}")


def main():
    started = datetime.now()
    log.info("")
    log.info("#" * 60)
    log.info(f"# Taco Bell Dashboard Auto-Update — {started.strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("#" * 60)

    # --skip-if-fresh: afternoon RETRY task support. If the morning run already
    # succeeded (latest_date is yesterday/today), exit quietly.
    if '--skip-if-fresh' in sys.argv:
        try:
            json_path = os.path.join(CONFIG["repo_path"], "data.json")
            with open(json_path, 'r', encoding='utf-8') as f:
                latest = (json.load(f).get('meta', {}) or {}).get('latest_date')
            if latest:
                lag_days = (datetime.now().date() - datetime.strptime(latest, '%Y-%m-%d').date()).days
                if lag_days <= 1:
                    log.info(f"Data is fresh (latest_date={latest}, lag={lag_days}d) — retry not needed. Exiting.")
                    return
                log.info(f"Data is STALE (latest_date={latest}, lag={lag_days}d) — morning run failed, retrying now...")
        except Exception as e:
            log.warning(f"--skip-if-fresh check failed ({e}) — proceeding with full run.")

    try:
        # 1. Get attachment(s) from Outlook
        xlsx_path, hourly_path = get_latest_attachment()

        # 2. Parse it
        sales_data = parse_sales_analysis(xlsx_path)

        # 2b. Parse hourly file (optional)
        hourly_data = parse_hourly_data(hourly_path)

        # 3. Build JSON (sales + budget + optional hourly)
        budget_path = CONFIG["budget_source_xlsx"]
        if Path(budget_path).exists():
            budget = load_budget(budget_path)
        else:
            log.info(f"STEP 3a: Budget file not found at {budget_path} - skipping (no budget data in dashboard)")
            budget = {}
        json_data = build_data_json(sales_data, budget, hourly_data=hourly_data)

        # 4. Write + push (with retries)
        git_push_data_json(json_data)

        # 5. Verify the deployment actually went live on GitHub Pages
        verify_deployment(json_data)

        elapsed = (datetime.now() - started).total_seconds()
        log.info("")
        log.info(f"✓ DONE in {elapsed:.1f}s")
        log.info("")

    except SystemExit:
        raise
    except Exception as e:
        log.exception(f"UNEXPECTED ERROR: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
