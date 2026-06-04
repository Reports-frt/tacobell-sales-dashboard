"""
Taco Bell Food Cost — Data Parser
============================

Reads the source food cost files and produces food_data.json
for the Food Cost Hub dashboard.

Source files expected:
  - FoodCost.xlsx          : Main historical data (Data sheet)
  - CategoriesFC.xlsx      : Ingredient code → category mapping
  - 2026_ΚΟΣΤΟΛΟΓΗΣΗ.xlsx   : Monthly COGS summary (used for cross-validation)
  - ΦΥΡΕΣ_MM_YYYY.xlsx      : Latest month waste analysis (optional)

Run:
    python build_food_data.py [--source-dir DIR] [--output-dir DIR]
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
from openpyxl import load_workbook


# =====================================================================
# CONFIG
# =====================================================================
KFC_STORES = [
    'KFC ΓΛΥΦΑΔΑ', 'KFC ΠΕΙΡΑΙΑΣ', 'KFC ΡΕΝΤΗΣ', 'KFC ΠΑΓΚΡΑΤΙ', 'KFC ΜΑΡΟΥΣΙ',
    'KFC COSMOS', 'KFC ESCAPE ΙΛΙΟΝ', 'KFC ΚΗΦΙΣΙΑΣ', 'KFC- RIVER WEST',
    'KFC- ΘΩΝ (Αμπελοκηποι)', 'KFC METRO MALL', 'KFC ΣΥΝΤΑΓΜΑ',
    'KFC Smart Park', 'KFC One Salonica', 'KFC Αριστοτέλους',
    'KFC N. ΣΜΥΡΝΗ', 'KFC ΧΑΛΑΝΔΡΙ', 'KFC Ν.ΦΙΛΑΔΕΛΦΕΙΑ',
    'KFC-ΓΕΡΑΚΑΣ', 'KFC-ΜΕΤΑΜΟΡΦΩΣΗ', 'KFC-ΠΑΤΡΑ', 'KFC-ΛΑΡΙΣΑ'
]

# Map FoodCost.xlsx names → canonical short names matching sales dashboard
STORE_NAME_MAP = {
    'KFC ΓΛΥΦΑΔΑ':              'ΓΛΥΦΑΔΑ',
    'KFC ΠΕΙΡΑΙΑΣ':             'ΠΕΙΡΑΙΑΣ',
    'KFC ΡΕΝΤΗΣ':               'ΡΕΝΤΗΣ',
    'KFC ΠΑΓΚΡΑΤΙ':             'ΠΑΓΚΡΑΤΙ',
    'KFC ΜΑΡΟΥΣΙ':              'ΜΑΡΟΥΣΙ',
    'KFC COSMOS':               'COSMOS',
    'KFC ESCAPE ΙΛΙΟΝ':         'ESCAPE',
    'KFC ΚΗΦΙΣΙΑΣ':             'ΚΗΦΙΣΙΑ',
    'KFC- RIVER WEST':          'RIVER',
    'KFC- ΘΩΝ (Αμπελοκηποι)':   'ΘΩΝ',
    'KFC METRO MALL':           'METROMALL',
    'KFC ΣΥΝΤΑΓΜΑ':             'ΣΥΝΤΑΓΜΑ',
    'KFC Smart Park':           'SMART PARK',
    'KFC One Salonica':         'ONE SALONICA',
    'KFC Αριστοτέλους':         'ΑΡΙΣΤΟΤΕΛΟΥΣ',
    'KFC N. ΣΜΥΡΝΗ':            'Ν. ΣΜΥΡΝΗ',
    'KFC ΧΑΛΑΝΔΡΙ':             'ΧΑΛΑΝΔΡΙ',
    'KFC Ν.ΦΙΛΑΔΕΛΦΕΙΑ':        'ΦΙΛΑΔΕΛΦΕΙΑ',
    'KFC-ΓΕΡΑΚΑΣ':              'ΓΕΡΑΚΑΣ',
    'KFC-ΜΕΤΑΜΟΡΦΩΣΗ':          'ΜΕΤΑΜΟΡΦΩΣΗ',
    'KFC-ΠΑΤΡΑ':                'ΠΑΤΡΑ',
    'KFC-ΛΑΡΙΣΑ':               'ΛΑΡΙΣΑ',
}

# Inventory Posting Group classification
# "Pure food" = ingredients that go INTO menu items
# "Total cost" = all of the above + packaging, services, etc.
PURE_FOOD_GROUPS = {'ΑΥΛΕΣ', 'ΠΡΟΙΟΝΤΑ', 'ΕΜΠΟΡΕΥΜΑΤ'}
ALL_FOOD_GROUPS = {'ΑΥΛΕΣ', 'ΠΡΟΙΟΝΤΑ', 'ΕΜΠΟΡΕΥΜΑΤ',
                   'ΥΠΗΡΕΣΙΕΣ', 'ΑΝΑΛΩΣΙΜΑ', 'ΑΝΤΑΛ/ΚΑ', 'SMALLWARES'}


# =====================================================================
# PARSING
# =====================================================================
def load_categories(path):
    """Read CategoriesFC.xlsx → dict {code: category}."""
    print(f"  Loading categories from {path.name}...")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Cat']
    cats = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            cats[str(row[0]).strip()] = str(row[2]).strip()
    wb.close()
    print(f"    → {len(cats)} ingredient codes mapped to categories")
    return cats


def load_foodcost_data(path):
    """Read FoodCost.xlsx 'Data' sheet → DataFrame, cleaned and KFC-only."""
    print(f"  Loading historical data from {path.name} ({path.stat().st_size / 1024 / 1024:.1f} MB)...")
    df = pd.read_excel(path, sheet_name='Data', header=2)
    print(f"    → {len(df):,} raw rows loaded")
    
    # Rename columns to short, lowercase keys
    df = df.rename(columns={
        'Έτος': 'year',
        'Μήνας': 'month',
        'Code': 'code',
        'Description': 'desc',
        'Base Unit of Measure': 'unit',
        'Inventory Posting Group': 'group',
        'Reason Code': 'reason',
        'Location Code': 'loc_code',
        'Location Name': 'loc_name',
        'Ideal(Quantity)': 'ideal_qty',
        'Other(Quantity)': 'other_qty',
        'Total(Quantity)': 'total_qty',
        'Ideal(Cost Amount Actual)': 'ideal_cost',
        'Other(Cost Amount Actual)': 'other_cost',
        'Total(Cost Amount Actual)': 'total_cost',
        'Category': 'category_existing',
    })
    
    # Filter to KFC stores only
    before = len(df)
    df = df[df['loc_name'].isin(KFC_STORES)].copy()
    print(f"    → {len(df):,} rows after filtering to {len(KFC_STORES)} KFC stores (dropped {before - len(df):,})")
    
    # Cost values are negative in source (they represent consumption/outflow).
    # Flip sign so positive = consumption €.
    for col in ['ideal_qty','other_qty','total_qty','ideal_cost','other_cost','total_cost']:
        df[col] = -df[col].fillna(0)
    
    # Map store names to canonical short names
    df['store'] = df['loc_name'].map(STORE_NAME_MAP)
    
    return df


def aggregate_monthly(df, categories):
    """
    Build the monthly aggregate structure.
    
    Returns dict {(year, month, store): {pure_food, total_cost, breakdown_by_cat, ...}}
    """
    print("  Aggregating monthly data...")
    
    # Add category from lookup (override existing if any)
    df['category'] = df['code'].map(categories).fillna(df['group'])
    
    # Mark pure-food vs other
    df['is_pure_food'] = df['group'].isin(PURE_FOOD_GROUPS)
    
    # Monthly rollups
    monthly = []
    for (year, month, store), grp in df.groupby(['year', 'month', 'store']):
        pure = grp[grp['is_pure_food']]
        total = grp[grp['group'].isin(ALL_FOOD_GROUPS)]
        
        rec = {
            'year': int(year),
            'month': int(month),
            'store': store,
            
            # Two flavors of food cost (toggleable in dashboard)
            'pure_food': {
                'ideal_cost':  round(float(pure['ideal_cost'].sum()), 2),
                'other_cost':  round(float(pure['other_cost'].sum()), 2),
                'total_cost':  round(float(pure['total_cost'].sum()), 2),
                'ideal_qty':   round(float(pure['ideal_qty'].sum()), 2),
                'total_qty':   round(float(pure['total_qty'].sum()), 2),
            },
            'total': {
                'ideal_cost':  round(float(total['ideal_cost'].sum()), 2),
                'other_cost':  round(float(total['other_cost'].sum()), 2),
                'total_cost':  round(float(total['total_cost'].sum()), 2),
                'ideal_qty':   round(float(total['ideal_qty'].sum()), 2),
                'total_qty':   round(float(total['total_qty'].sum()), 2),
            },
            
            # Waste breakdown (Reason Code = ΦΥΡΑ)
            'waste': {
                'cost': round(float(total[total['reason'] == 'ΦΥΡΑ']['total_cost'].sum()), 2),
                'qty':  round(float(total[total['reason'] == 'ΦΥΡΑ']['total_qty'].sum()), 2),
            },
            
            # Reason breakdown (cost by reason, food relevant)
            'by_reason': {
                str(r) if r is not None and pd.notna(r) else 'NORMAL':
                    round(float(total[total['reason'].fillna('NORMAL').replace({None: 'NORMAL'}) == 
                                       (r if pd.notna(r) else 'NORMAL')]['total_cost'].sum()), 2)
                for r in total['reason'].fillna('NORMAL').unique()
            },
            
            # Category breakdown (pure food)
            'by_category': pure.groupby('category').agg(
                ideal_cost=('ideal_cost', 'sum'),
                other_cost=('other_cost', 'sum'),
                total_cost=('total_cost', 'sum'),
            ).round(2).to_dict(orient='index'),
        }
        monthly.append(rec)
    
    print(f"    → {len(monthly)} (year × month × store) records")
    return monthly


def aggregate_top_variances(df, categories, year, month):
    """
    For a given month, compute top-variance ingredients per store.
    Returns dict {store: [{code, desc, category, ideal_cost, other_cost, total_cost}, ...]}
    """
    df['category'] = df['code'].map(categories).fillna(df['group'])
    pure = df[df['group'].isin(PURE_FOOD_GROUPS)].copy()
    
    target = pure[(pure['year'] == year) & (pure['month'] == month)].copy()
    
    result = {}
    for store, grp in target.groupby('store'):
        items = grp.groupby(['code','desc','category']).agg(
            ideal_cost=('ideal_cost', 'sum'),
            other_cost=('other_cost', 'sum'),
            total_cost=('total_cost', 'sum'),
            total_qty=('total_qty', 'sum'),
        ).reset_index()
        items['abs_other'] = items['other_cost'].abs()
        # Top 20 by absolute variance
        top = items.nlargest(20, 'abs_other')
        result[store] = [
            {
                'code': str(row['code']),
                'desc': str(row['desc']),
                'category': str(row['category']),
                'ideal_cost': round(float(row['ideal_cost']), 2),
                'other_cost': round(float(row['other_cost']), 2),
                'total_cost': round(float(row['total_cost']), 2),
                'total_qty': round(float(row['total_qty']), 2),
            }
            for _, row in top.iterrows()
        ]
    
    return result


def load_kostologisi(path):
    """
    Read 2026_ΚΟΣΤΟΛΟΓΗΣΗ.xlsx → for cross-validation.
    Returns dict {(year, month, store): {αρχικο, αγορες, τελικο, COGS}}
    """
    print(f"  Loading ΚΟΣΤΟΛΟΓΗΣΗ summary from {path.name}...")
    wb = load_workbook(path, data_only=True)
    ws = wb['ΚΟΣΤΟΛΟΓΗΣΗ']
    
    # Detect month column starts (row 2 has period headers like "Food Cost _ JAN 2026")
    period_starts = []
    for cell in ws[2]:
        if cell.value and 'Food Cost' in str(cell.value) and 'ALL' not in str(cell.value):
            period_starts.append((cell.column, str(cell.value)))
    
    # Parse month from "Food Cost _JAN 2026" or "Food Cost _ FEB 2026"
    MONTHS = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,
              'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
    
    result = {}
    for col_start, header in period_starts:
        # Extract month + year from header
        parts = header.replace('Food Cost', '').replace('_', '').strip().split()
        month_str = parts[0].strip()
        year = int(parts[-1])
        if month_str not in MONTHS:
            continue
        month = MONTHS[month_str]
        
        # Read 22 store rows (4-25)
        for r in range(4, 26):
            store_name = ws.cell(row=r, column=col_start).value  # "[201] KFC-ΓΛΥΦΑΔΑ"
            cogs = ws.cell(row=r, column=col_start + 5).value  # Κόστος Πωληθέντων
            if store_name and cogs is not None:
                result[(year, month, str(store_name).strip())] = {
                    'cogs': float(cogs)
                }
    wb.close()
    print(f"    → {len(result)} period × store records loaded")
    return result



# =====================================================================
# TARGIT XLSX CLEANER
# =====================================================================
def _read_targit_xlsx(path, sheet_name='Object1'):
    """
    Read xlsx exported by Targit, working around two known issues:
      1. Non-standard `defaultColWidthPt` attribute that breaks openpyxl
      2. Header labels split across rows (some in row 0, others in row 1)
    """
    import zipfile, re, tempfile, os
    
    # Step 1: clean the xlsx into a temp file
    tmpdir = tempfile.mkdtemp()
    cleaned_path = os.path.join(tmpdir, 'cleaned.xlsx')
    try:
        with zipfile.ZipFile(str(path), 'r') as zin:
            with zipfile.ZipFile(cleaned_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    content = zin.read(item)
                    if item.endswith('.xml'):
                        text = content.decode('utf-8', errors='ignore')
                        text = re.sub(r'\s+defaultColWidthPt="[^"]*"', '', text)
                        text = re.sub(r'\s+defaultRowHeightPt="[^"]*"', '', text)
                        content = text.encode('utf-8')
                    zout.writestr(item, content)
    except zipfile.BadZipFile:
        # Already a clean file
        cleaned_path = str(path)
    
    # Step 2: read raw and detect header row
    df_raw = pd.read_excel(cleaned_path, sheet_name=sheet_name, header=None)
    
    # Heuristic: header row is the one where most cells are non-null and contain 
    # text (not numbers). Check first 3 rows.
    header_row = 0
    for r in range(min(3, len(df_raw))):
        row_vals = df_raw.iloc[r].dropna()
        if len(row_vals) >= len(df_raw.columns) * 0.7:  # 70%+ filled
            text_count = sum(1 for v in row_vals if isinstance(v, str))
            if text_count >= len(row_vals) * 0.7:
                header_row = r
                break
    
    # Build merged header: take row 0 and row 1, prefer non-null
    if header_row == 0 and len(df_raw) > 1:
        # Check if row 0 has gaps that row 1 fills
        h0 = df_raw.iloc[0].tolist()
        h1 = df_raw.iloc[1].tolist()
        merged = []
        for c0, c1 in zip(h0, h1):
            if pd.notna(c0) and str(c0).strip():
                merged.append(str(c0).strip())
            elif pd.notna(c1) and str(c1).strip():
                merged.append(str(c1).strip())
            else:
                merged.append(f'col_{len(merged)}')
        df = df_raw.iloc[2:].copy()
        df.columns = merged
        df = df.reset_index(drop=True)
    else:
        df = pd.read_excel(cleaned_path, sheet_name=sheet_name, header=header_row)
    
    return df


def _safe_float(x):
    """Convert to float, returning 0 for non-numeric values like 'Μαθηματικό σφάλμα'."""
    if x is None:
        return 0.0
    try:
        if pd.isna(x):
            return 0.0
    except (TypeError, ValueError):
        pass
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0

def _safe_int(x):
    return int(_safe_float(x))

def load_meikto_kerdos(path):
    """
    Read Μεικτό_κέρδος_KFC_new.xlsx
    Object1 = per product × hour breakdown
    Object2 = per store × sales type breakdown
    
    Returns dict with:
      - products: list of product-level records (chain-wide totals)
      - stores: list of store × sales-type records  
      - hourly: { hour: { product: {qty, sales, fc, fc_pct, rating} } }
    """
    print(f"  Loading product profitability from {path.name}...")
    
    # Object1: per product × hour
    df1 = _read_targit_xlsx(path, sheet_name='Object1')
    df1 = df1.rename(columns={
        df1.columns[0]: 'hour',
        'ITEM Product Group Description': 'category',
        'ITEM Description': 'product',
        'Quantity': 'qty',
        '% Qty Contrib.': 'qty_share',
        'Net Sales': 'sales',
        '% NS Contrib.': 'sales_share',
        'Food Cost': 'fc',
        '% FC Contrib.': 'fc_share',
        'Average Sale': 'avg_sale',
        'Average FC': 'avg_fc',
        'Average Gross Margin (GM)': 'avg_gm',
        '% FoodCost [FC/NS]': 'fc_pct',
        '% GM [GM/NS]': 'gm_pct',
        'Rating': 'rating',
    })
    
    # Aggregate by product across all hours (since each product appears in multiple hour rows)
    # Skip rows where category or product is "Συνολικά" (those are aggregations)
    df_prod = df1[
        (df1['category'].astype(str).str.strip() != 'Συνολικά') &
        (df1['product'].astype(str).str.strip() != 'Συνολικά') &
        (df1['hour'].astype(str).str.strip() != 'Συνολικά') &
        (df1['qty'].notna())
    ].copy()
    
    # Convert numeric cols to numeric (will turn 'Μαθηματικό σφάλμα' into NaN)
    df_prod['qty'] = pd.to_numeric(df_prod['qty'], errors='coerce').fillna(0)
    df_prod['sales'] = pd.to_numeric(df_prod['sales'], errors='coerce').fillna(0)
    df_prod['fc'] = pd.to_numeric(df_prod['fc'], errors='coerce').fillna(0)
    
    # Group and sum
    agg = df_prod.groupby(['category','product']).agg(
        qty=('qty','sum'),
        sales=('sales','sum'),
        fc=('fc','sum'),
    ).reset_index()
    
    products = []
    for _, r in agg.iterrows():
        qty = _safe_float(r['qty'])
        sales = _safe_float(r['sales'])
        fc = _safe_float(r['fc'])
        if qty < 1:
            continue
        # Compute derived metrics from totals (safe — protects against zero/None)
        fc_pct = fc/sales if sales > 0 else 0
        avg_sale = sales/qty if qty > 0 else 0
        avg_fc = fc/qty if qty > 0 else 0
        avg_gm = avg_sale - avg_fc
        # Rating = % qty contribution chain-wide
        # Will compute after we have total
        products.append({
            'category': str(r['category']),
            'product': str(r['product']),
            'qty': _safe_int(r['qty']),
            'sales': round(_safe_float(r['sales']), 2),
            'fc': round(_safe_float(r['fc']), 2),
            'fc_pct': round(_safe_float(fc_pct), 4),
            'avg_sale': round(_safe_float(avg_sale), 2),
            'avg_gm': round(_safe_float(avg_gm), 2),
            'rating': 0,  # filled below
        })
    
    # Compute rating (% qty contribution)
    total_qty = sum(p['qty'] for p in products)
    if total_qty > 0:
        for p in products:
            p['rating'] = round(p['qty'] / total_qty, 6)
    
    print(f"    → {len(products)} unique products with profitability data")
    
    # Hourly breakdown — only category-level totals to keep size manageable
    hourly = {}
    for _, r in df1.iterrows():
        hour = str(r['hour']).strip()
        if hour == 'Συνολικά':
            continue
        cat = str(r['category']).strip()
        prod = str(r['product']).strip()
        # Only category aggregations (not individual products to limit size)
        if prod != 'Συνολικά':
            continue
        if cat == 'Συνολικά':
            # Hour total
            key = '_total'
        else:
            key = cat
        if hour not in hourly:
            hourly[hour] = {}
        hourly[hour][key] = {
            'qty': _safe_int(r.get('qty')),
            'sales': round(_safe_float(r.get('sales')), 2),
            'fc': round(_safe_float(r.get('fc')), 2),
            'fc_pct': round(_safe_float(r.get('fc_pct')), 4),
            'rating': round(_safe_float(r.get('rating')), 6),
        }
    print(f"    → {len(hourly)} hours of category-level data")
    
    # Object2: per store × sales type
    df2 = _read_targit_xlsx(path, sheet_name='Object2')
    df2.columns = [str(c).strip() for c in df2.columns]
    
    rename_map = {
        'BRAND': 'brand',
        'Name': 'store_name',
        'Hour Of Day': 'hour',
        'Sales Type Description': 'sales_type',
        'Quantity': 'qty',
        'Net Sales': 'sales',
        'Food Cost': 'fc',
        'Average Sale': 'avg_sale',
        'Average FC': 'avg_fc',
        'Average Gross Margin (GM)': 'avg_gm',
        '% FoodCost [FC/NS]': 'fc_pct',
        '% GM [GM/NS]': 'gm_pct',
    }
    df2 = df2.rename(columns={k: v for k, v in rename_map.items() if k in df2.columns})
    
    has_store = 'store_name' in df2.columns
    has_hour = 'hour' in df2.columns
    
    # Map raw store names to canonical
    STORE_MAP_OBJ2 = {
        'KFC -  ΘΩΝ (ΑΜΠΕΛΟΚΗΠΟΙ)': 'ΘΩΝ',
        'KFC - COSMOS': 'COSMOS',
        'KFC - ESCAPE (ΙΛΙΟΝ)': 'ESCAPE',
        'KFC - MALL (ΜΑΡΟΥΣΙ)': 'ΜΑΡΟΥΣΙ',
        'KFC - METRO MALL': 'METROMALL',
        'KFC - ONE SALONICA': 'ONE SALONICA',
        'KFC - RIVER WEST': 'RIVER',
        'KFC - SMART PARK': 'SMART PARK',
        'KFC - VILLAGE PARK (ΡΕΝΤΗΣ)': 'ΡΕΝΤΗΣ',
        'KFC - ΑΡΙΣΤΟΤΕΛΟΥΣ': 'ΑΡΙΣΤΟΤΕΛΟΥΣ',
        'KFC - ΓΕΡΑΚΑ': 'ΓΕΡΑΚΑΣ',
        'KFC - ΓΛΥΦΑΔΑ': 'ΓΛΥΦΑΔΑ',
        'KFC - ΚΗΦIΣΙΑ': 'ΚΗΦΙΣΙΑ',
        'KFC - ΛΑΡΙΣΑ': 'ΛΑΡΙΣΑ',
        'KFC - ΜΕΤΑΜΟΡΦΩΣΗ': 'ΜΕΤΑΜΟΡΦΩΣΗ',
        'KFC - Ν. ΣΜΥΡΝΗΣ': 'Ν. ΣΜΥΡΝΗ',
        'KFC - ΝΕΑ ΦΙΛΑΔΕΛΦΕΙΑ': 'ΦΙΛΑΔΕΛΦΕΙΑ',
        'KFC - ΠΑΓΚΡΑΤΙ': 'ΠΑΓΚΡΑΤΙ',
        'KFC - ΠΑΤΡΑ': 'ΠΑΤΡΑ',
        'KFC - ΠΕΙΡΑΙΑ': 'ΠΕΙΡΑΙΑΣ',
        'KFC - ΣΥΝΤΑΓΜΑ': 'ΣΥΝΤΑΓΜΑ',
        'KFC - ΧΑΛΑΝΔΡΙ': 'ΧΑΛΑΝΔΡΙ',
    }
    
    # Build comprehensive structures:
    # 1. Per-store totals (hour=Συνολικά, sales_type=Συνολικά)
    # 2. Per-store hourly (hour!=Συνολικά, sales_type=Συνολικά)
    # 3. Per-store sales-type (hour=Συνολικά, sales_type!=Συνολικά)
    
    store_totals = []          # [{store, qty, sales, fc, fc_pct}]
    store_hourly = {}          # {store: {hour: {qty, sales, fc, fc_pct}}}
    store_salestype = {}       # {store: {sales_type: {qty, sales, fc, fc_pct}}}
    store_hour_salestype = {}  # {store: {hour: {sales_type: {...}}}}
    
    def make_record(r):
        return {
            'qty': _safe_int(r.get('qty')),
            'sales': round(_safe_float(r.get('sales')), 2),
            'fc': round(_safe_float(r.get('fc')), 2),
            'fc_pct': round(_safe_float(r.get('fc_pct')), 4),
            'avg_sale': round(_safe_float(r.get('avg_sale')), 2),
            'avg_gm': round(_safe_float(r.get('avg_gm')), 2),
        }
    
    for _, r in df2.iterrows():
        store_raw = str(r.get('store_name', '')).strip() if has_store else ''
        hour = str(r.get('hour', '')).strip() if has_hour else 'Συνολικά'
        st = str(r.get('sales_type', '')).strip()
        
        if pd.isna(r.get('qty')):
            continue
        if store_raw in ('', 'nan', 'Συνολικά'):
            continue  # Skip aggregate or missing store
        
        store = STORE_MAP_OBJ2.get(store_raw, store_raw)
        rec = make_record(r)
        
        # Categorize this row
        is_hour_total = (hour == 'Συνολικά' or hour == 'nan')
        is_st_total = (st == 'Συνολικά' or st == 'nan')
        
        if is_hour_total and is_st_total:
            # Store grand total
            store_totals.append({'store': store, **rec})
        elif is_hour_total and not is_st_total:
            # Per-store sales type breakdown
            if store not in store_salestype:
                store_salestype[store] = {}
            store_salestype[store][st] = rec
        elif not is_hour_total and is_st_total:
            # Per-store hourly
            if store not in store_hourly:
                store_hourly[store] = {}
            store_hourly[store][hour] = rec
        else:
            # Full hour × store × sales type
            if store not in store_hour_salestype:
                store_hour_salestype[store] = {}
            if hour not in store_hour_salestype[store]:
                store_hour_salestype[store][hour] = {}
            store_hour_salestype[store][hour][st] = rec
    
    print(f"    → store totals: {len(store_totals)}")
    print(f"    → store hourly: {len(store_hourly)} stores × ~{sum(len(v) for v in store_hourly.values())//max(1,len(store_hourly))} hours each")
    print(f"    → store sales types: {len(store_salestype)} stores")
    
    return {
        'products': products,
        'hourly_category': hourly,            # chain-wide hour × category
        'store_totals': store_totals,         # per-store totals
        'store_hourly': store_hourly,         # store × hour
        'store_salestype': store_salestype,   # store × sales type
    }


def load_alc_combos(path, kind='alc'):
    """
    Read ALC or Combos hourly file.
    Object1 has multi-index columns: (store, hour) for ALC or (store, hour×2) for Combos.
    
    For ALC: Each cell = quantity for store × product × hour
    For Combos: Each pair of cells = (Net Sales €, Receipt count) for store × product × hour
    
    Returns:
      ALC: {store: {product: {hour: qty}}}
      Combos: {store: {product: {hour: {sales: €, receipts: n}}}}
    """
    print(f"  Loading {kind.upper()} hourly data from {path.name}...")
    
    # Read header rows to determine structure
    raw = pd.read_excel(path, sheet_name='Object1', header=None, nrows=5)
    
    # Find the description columns: rows where row 0/1 are NaN/Unnamed and contain product info
    # Identify row index where headers transition to data
    # Strategy: find row where col 0 has a real product string (not NaN/Συνολικά)
    
    # First two rows are typically header (store name) and (hour or metric)
    # For ALC: row 0 = stores, row 1 = hours, row 2 = column metric labels (often missing)
    # For Combos: row 0 = stores, row 1 = hours+suffix, row 2 = metric labels (Καθαρή Αξία | Line Receipt)
    
    # Re-read with header=[0,1] then check for metric row
    df = pd.read_excel(path, sheet_name='Object1', header=[0, 1])
    
    # Determine description columns: those with both header levels containing 'Unnamed' or NaN
    desc_col_count = 0
    for col in df.columns:
        l0 = str(col[0])
        if 'Unnamed' in l0:
            desc_col_count += 1
        else:
            break
    
    # Determine metric row: read first data row to see if it contains metric labels
    is_combos_with_metrics = False
    if len(df) > 0:
        first_row = df.iloc[0]
        # Check if first row contains 'Καθαρή Αξία' or 'Quantity' indicators
        sample = str(first_row.iloc[desc_col_count] if len(first_row) > desc_col_count else '')
        if 'Αξία' in sample or 'Quantity' in sample or 'Receipt' in sample:
            is_combos_with_metrics = True
    
    # Use desc cols to extract product info
    # Identify store-hour columns
    store_hour_cols = {}  # {store: [(col_idx, hour, metric_idx)]}
    for i in range(desc_col_count, len(df.columns)):
        col = df.columns[i]
        store = str(col[0]).strip()
        hour = str(col[1]).strip()
        if 'Unnamed' in store or store == '':
            continue
        if store not in store_hour_cols:
            store_hour_cols[store] = []
        store_hour_cols[store].append((i, hour))
    
    print(f"    → {len(store_hour_cols)} stores detected (desc cols: {desc_col_count}, has metrics row: {is_combos_with_metrics})")
    
    # Skip the metrics row if it exists
    data_start = 1 if is_combos_with_metrics else 0
    
    result = {}
    products_seen = set()
    
    for ridx in range(data_start, len(df)):
        row = df.iloc[ridx]
        # Build product key from desc cols
        parts = []
        for di in range(desc_col_count):
            v = row.iloc[di]
            if pd.notna(v) and str(v).strip() not in ('nan', 'Συνολικά', ''):
                parts.append(str(v).strip())
        if not parts:
            continue
        product_key = ' | '.join(parts)
        # Skip if all parts are "Συνολικά"
        if all(p == 'Συνολικά' for p in parts):
            continue
        products_seen.add(product_key)
        
        for store, cols in store_hour_cols.items():
            if store not in result:
                result[store] = {}
            if product_key not in result[store]:
                result[store][product_key] = {}
            
            # For combos with paired columns, alternate Sales/Receipts
            if is_combos_with_metrics and kind == 'combos':
                # Pair up consecutive cols (sales €, receipts)
                for j in range(0, len(cols) - 1, 2):
                    col_idx_a, hour_a = cols[j]
                    col_idx_b, hour_b = cols[j+1]
                    sales = row.iloc[col_idx_a]
                    receipts = row.iloc[col_idx_b]
                    # Hour from "07:00" or "07:00.1" — strip suffix
                    hour_clean = hour_a.split('.')[0] if '.' in hour_a else hour_a
                    if pd.notna(sales) and float(sales) != 0:
                        result[store][product_key][hour_clean] = {
                            'sales': round(float(sales), 2),
                            'receipts': int(receipts) if pd.notna(receipts) else 0,
                        }
            else:
                # Simple: each hour col = qty
                for col_idx, hour in cols:
                    val = row.iloc[col_idx]
                    if pd.notna(val) and float(val) != 0:
                        result[store][product_key][hour] = float(val)
    
    # Filter out aggregate "Συνολικά" pseudo-store
    result = {k: v for k, v in result.items() if k.strip() != 'Συνολικά'}
    
    # Map raw store names to canonical short names where possible
    ALC_STORE_MAP = {
        'KFC -  ΘΩΝ (ΑΜΠΕΛΟΚΗΠΟΙ)': 'ΘΩΝ',
        'KFC - COSMOS': 'COSMOS',
        'KFC - ESCAPE (ΙΛΙΟΝ)': 'ESCAPE',
        'KFC - MALL (ΜΑΡΟΥΣΙ)': 'ΜΑΡΟΥΣΙ',
        'KFC - METRO MALL': 'METROMALL',
        'KFC - ONE SALONICA': 'ONE SALONICA',
        'KFC - RIVER WEST': 'RIVER',
        'KFC - SMART PARK': 'SMART PARK',
        'KFC - VILLAGE PARK (ΡΕΝΤΗΣ)': 'ΡΕΝΤΗΣ',
        'KFC - ΑΡΙΣΤΟΤΕΛΟΥΣ': 'ΑΡΙΣΤΟΤΕΛΟΥΣ',
        'KFC - ΓΕΡΑΚΑ': 'ΓΕΡΑΚΑΣ',
        'KFC - ΓΛΥΦΑΔΑ': 'ΓΛΥΦΑΔΑ',
        'KFC - ΚΗΦIΣΙΑ': 'ΚΗΦΙΣΙΑ',
        'KFC - ΛΑΡΙΣΑ': 'ΛΑΡΙΣΑ',
        'KFC - ΜΕΤΑΜΟΡΦΩΣΗ': 'ΜΕΤΑΜΟΡΦΩΣΗ',
        'KFC - Ν. ΣΜΥΡΝΗΣ': 'Ν. ΣΜΥΡΝΗ',
        'KFC - ΝΕΑ ΦΙΛΑΔΕΛΦΕΙΑ': 'ΦΙΛΑΔΕΛΦΕΙΑ',
        'KFC - ΠΑΓΚΡΑΤΙ': 'ΠΑΓΚΡΑΤΙ',
        'KFC - ΠΑΤΡΑ': 'ΠΑΤΡΑ',
        'KFC - ΠΕΙΡΑΙΑ': 'ΠΕΙΡΑΙΑΣ',
        'KFC - ΣΥΝΤΑΓΜΑ': 'ΣΥΝΤΑΓΜΑ',
        'KFC - ΧΑΛΑΝΔΡΙ': 'ΧΑΛΑΝΔΡΙ',
    }
    result_mapped = {}
    for raw_store, products_dict in result.items():
        canonical = ALC_STORE_MAP.get(raw_store, raw_store)
        result_mapped[canonical] = products_dict
    result = result_mapped
    
    # Strip "Συνολικά | " prefix from product keys (it pollutes mapped names)
    for store in result:
        cleaned = {}
        for prod_key, hours_data in result[store].items():
            # Remove leading "Συνολικά | " if present
            new_key = prod_key.replace('Συνολικά | ', '').strip()
            cleaned[new_key] = hours_data
        result[store] = cleaned
    
    # Compute total records
    total_records = sum(
        sum(len(v) for v in store_dict.values()) 
        for store_dict in result.values()
    )
    print(f"    → {len(products_seen)} unique products, ~{total_records} store-product-hour records")
    print(f"    → {len(result)} stores after canonical mapping")
    return {
        'stores': result,
        'product_count': len(products_seen),
    }


# =====================================================================
# BUILD JSON
# =====================================================================

# =====================================================================
# QUANTITY PER ITEM (sheet from FoodCost.xlsx)
# =====================================================================
def load_quantity_per_item(path):
    """
    Read the 'Quantity Per Item' sheet from FoodCost.xlsx.
    
    Format:
      Row 1: Έτος | 2025
      Row 2: Μήνας | 12
      Row 6: Code | Description | Unit | Ideal Qty | Other Qty | Total Qty | (cost?) | waste% | ...
      Rows 7+: ingredients (677 rows)
    
    Returns dict: { 'period': 'YYYY-MM', 'items': [...] }
    """
    try:
        # Get the period from rows 1-2
        meta = pd.read_excel(path, sheet_name='Quantity Per Item', header=None, nrows=3)
        year = None
        month = None
        for _, r in meta.iterrows():
            lbl = str(r.iloc[0] or '').strip()
            val = r.iloc[1] if len(r) > 1 else None
            if lbl == 'Έτος' and pd.notna(val):
                try:
                    year = int(val)
                except (ValueError, TypeError):
                    pass
            elif lbl == 'Μήνας' and pd.notna(val):
                try:
                    month = int(val)
                except (ValueError, TypeError):
                    pass
        
        period = f"{year:04d}-{month:02d}" if year and month else None
        
        # Read data starting at row 6 (header) — meaning data starts row 7
        df = pd.read_excel(path, sheet_name='Quantity Per Item', header=6)
        # Drop rows where Code is missing
        df = df[df['Code'].notna()].copy()
        df = df[df['Code'].astype(str).str.strip() != ''].copy()
        df = df[~df['Code'].astype(str).str.contains('Total', case=False, na=False)].copy()
        
        # Numeric conversion (safe)
        for c in ['Sum of Ideal(Quantity)', 'Sum of Other(Quantity)', 'Sum of Total(Quantity)']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        
        items = []
        for _, r in df.iterrows():
            code = str(r['Code']).strip()
            if not code or code.lower() == 'nan':
                continue
            ideal = float(r.get('Sum of Ideal(Quantity)', 0) or 0)
            other = float(r.get('Sum of Other(Quantity)', 0) or 0)
            total = float(r.get('Sum of Total(Quantity)', 0) or 0)
            # Compute waste% = abs(other / ideal) when ideal != 0
            # Negative ideal means consumed (good), positive other means waste/return
            waste_pct = 0
            if abs(ideal) > 0.01:
                waste_pct = abs(other / ideal)
            items.append({
                'code': code,
                'description': str(r.get('Description', '')).strip(),
                'unit': str(r.get('Base Unit of Measure', '')).strip(),
                'ideal_qty': round(ideal, 3),
                'other_qty': round(other, 3),
                'total_qty': round(total, 3),
                'waste_pct': round(waste_pct, 4),
            })
        
        # Sort by absolute total qty (largest first)
        items.sort(key=lambda x: abs(x['total_qty']), reverse=True)
        
        print(f"  Loaded {len(items)} ingredients from 'Quantity Per Item' sheet (period {period})")
        return {'period': period, 'items': items}
    except Exception as e:
        print(f"  Could not load 'Quantity Per Item' sheet: {e}")
        return None


# =====================================================================
# REAL-TIME METRICS (computed from daily Μεικτό κέρδος)
# =====================================================================
def compute_realtime_metrics(meikto_data, monthly_records, kfc_stores):
    """
    Build a 'real-time' snapshot from the latest Μεικτό κέρδος data.
    
    Returns dict with:
      - chain: { sales, fc, fc_pct, qty, n_stores }
      - per_store: list of { store, sales, fc, fc_pct, qty, vs_chain_pp }
      - hourly: list of { hour, sales, fc, fc_pct }  (skip dead hours 02-10)
      - top_fc: list of { product, category, sales, fc, fc_pct, contribution_pp }
      - vs_history: { last_month_avg_fc_pct, delta_pp }
    """
    if not meikto_data:
        return None
    
    rt = {}
    
    # ========== Chain-wide totals ==========
    products = meikto_data.get('products', [])
    chain_sales = sum(p.get('sales', 0) for p in products)
    chain_fc = sum(p.get('fc', 0) for p in products)
    chain_qty = sum(p.get('qty', 0) for p in products)
    chain_fc_pct = chain_fc / chain_sales if chain_sales else 0
    
    rt['chain'] = {
        'sales': round(chain_sales, 2),
        'fc': round(chain_fc, 2),
        'fc_pct': round(chain_fc_pct, 4),
        'qty': int(chain_qty),
        'n_products': len(products),
    }
    
    # ========== Per-store FC% ranking ==========
    store_totals = meikto_data.get('store_totals', [])
    store_hourly = meikto_data.get('store_hourly', {})
    
    # Build per-store totals by aggregating store_hourly (since store_totals may be empty)
    per_store_dict = {}
    for store_name, hours_data in store_hourly.items():
        s_sales = sum(h.get('sales', 0) for h in hours_data.values())
        s_fc = sum(h.get('fc', 0) for h in hours_data.values())
        s_qty = sum(h.get('qty', 0) for h in hours_data.values())
        per_store_dict[store_name] = {
            'sales': s_sales, 'fc': s_fc, 'qty': s_qty,
            'fc_pct': s_fc / s_sales if s_sales else 0,
        }
    
    # Convert to list, add chain comparison
    per_store_list = []
    for store, vals in per_store_dict.items():
        per_store_list.append({
            'store': store,
            'sales': round(vals['sales'], 2),
            'fc': round(vals['fc'], 2),
            'fc_pct': round(vals['fc_pct'], 4),
            'qty': int(vals['qty']),
            'vs_chain_pp': round((vals['fc_pct'] - chain_fc_pct) * 100, 2),
        })
    
    # Sort by FC% descending (worst first)
    per_store_list.sort(key=lambda r: r['fc_pct'], reverse=True)
    rt['per_store'] = per_store_list
    
    # ========== Hourly FC% profile (chain-wide) ==========
    hourly_cat = meikto_data.get('hourly_category', {})
    DEAD_HOURS = set([2, 3, 4, 5, 6, 7, 8, 9, 10])
    
    hourly_list = []
    for hour, cats in hourly_cat.items():
        # Hour can be '08:00' string or int 8
        try:
            if isinstance(hour, str) and ':' in hour:
                h_int = int(hour.split(':')[0])
            else:
                h_int = int(hour)
        except (ValueError, TypeError):
            continue
        if h_int in DEAD_HOURS:
            continue
        h_sales = sum(c.get('sales', 0) for c in cats.values())
        h_fc = sum(c.get('fc', 0) for c in cats.values())
        h_qty = sum(c.get('qty', 0) for c in cats.values())
        if h_sales <= 0:
            continue
        hourly_list.append({
            'hour': h_int,
            'sales': round(h_sales, 2),
            'fc': round(h_fc, 2),
            'fc_pct': round(h_fc / h_sales, 4),
            'qty': int(h_qty),
        })
    hourly_list.sort(key=lambda r: r['hour'])
    rt['hourly'] = hourly_list
    
    # ========== Top FC contributors (by absolute FC €) ==========
    real_products = [p for p in products if p.get('fc', 0) > 0 and p.get('sales', 0) > 0]
    real_products.sort(key=lambda p: p.get('fc', 0), reverse=True)
    
    top_fc = []
    for p in real_products[:20]:
        contrib_pp = (p.get('fc', 0) / chain_fc * 100) if chain_fc else 0
        top_fc.append({
            'product': p.get('product', ''),
            'category': p.get('category', ''),
            'sales': p.get('sales', 0),
            'fc': p.get('fc', 0),
            'fc_pct': p.get('fc_pct', 0),
            'qty': p.get('qty', 0),
            'contribution_pp': round(contrib_pp, 2),
        })
    rt['top_fc'] = top_fc
    
    # ========== Comparison vs last 3 months avg ==========
    if monthly_records:
        # Get the last 3 months from monthly data
        recent = sorted(monthly_records, key=lambda r: (r['year'], r['month']), reverse=True)
        recent_periods = set()
        for r in recent:
            key = (r['year'], r['month'])
            if len(recent_periods) >= 3:
                break
            recent_periods.add(key)
        
        recent_data = [r for r in monthly_records if (r['year'], r['month']) in recent_periods]
        if recent_data:
            avg_ideal = sum(r['pure_food']['ideal_cost'] for r in recent_data)
            avg_actual = sum(r['pure_food']['other_cost'] + r['pure_food']['ideal_cost'] for r in recent_data)
            # Estimate sales from ideal cost (assuming ideal is ~theoretical FC%)
            # Better: just compare ideal_cost to actual_cost ratio
            ideal_pct = avg_ideal / avg_actual if avg_actual else 0
            
            # Try a different angle: compare current FC% to last month historical
            rt['vs_history'] = {
                'note': 'Current period vs cumulative ISO 2025-2026 baseline',
                'baseline_label': f"Last 3 months avg (from FoodCost.xlsx)",
                'periods_compared': sorted([f"{y}-{m:02d}" for (y, m) in recent_periods], reverse=True),
            }
    
    return rt


def update_daily_archive(realtime, archive_dir: Path, today_str=None):
    """
    Save today's snapshot to daily_archive/YYYY-MM-DD.json.
    Then build a history series of {date, sales, fc, fc_pct, qty} from all archives.
    """
    import datetime as _dt
    if today_str is None:
        today_str = _dt.date.today().isoformat()
    
    archive_dir.mkdir(parents=True, exist_ok=True)
    
    # Snapshot today's data
    if realtime and 'chain' in realtime:
        snapshot_path = archive_dir / f"{today_str}.json"
        snapshot = {
            'date': today_str,
            'chain': realtime['chain'],
            'per_store_summary': [
                {'store': s['store'], 'sales': s['sales'], 'fc': s['fc'], 'fc_pct': s['fc_pct']}
                for s in realtime.get('per_store', [])
            ],
        }
        with open(snapshot_path, 'w', encoding='utf-8') as f:
            json.dump(snapshot, f, ensure_ascii=False, separators=(',', ':'))
        print(f"  → Saved daily snapshot: {snapshot_path.name}")
    
    # Load all archive snapshots and build history
    history = []
    if archive_dir.exists():
        for fp in sorted(archive_dir.glob('*.json')):
            try:
                with open(fp, 'r', encoding='utf-8') as f:
                    snap = json.load(f)
                history.append({
                    'date': snap['date'],
                    'sales': snap['chain']['sales'],
                    'fc': snap['chain']['fc'],
                    'fc_pct': snap['chain']['fc_pct'],
                    'qty': snap['chain']['qty'],
                })
            except Exception as e:
                print(f"  ⚠ Skipped corrupted archive {fp.name}: {e}")
                continue
    
    # Compute day-over-day deltas
    for i, h in enumerate(history):
        if i == 0:
            h['delta_sales'] = 0
            h['delta_fc_pp'] = 0
        else:
            prev = history[i-1]
            h['delta_sales'] = round(h['sales'] - prev['sales'], 2)
            h['delta_fc_pp'] = round((h['fc_pct'] - prev['fc_pct']) * 100, 2)
    
    print(f"  → Daily history: {len(history)} snapshots ({history[0]['date'] if history else 'none'} → {history[-1]['date'] if history else 'none'})")
    return history


def build_food_data_json(source_dir: Path, output_dir: Path):
    """Main pipeline: read sources, aggregate, write food_data.json."""
    print("=" * 60)
    print("Taco Bell Food Cost Hub — Data Build")
    print("=" * 60)
    print(f"Source: {source_dir}")
    print(f"Output: {output_dir}")
    print()
    
    # Step 1: Load lookups
    print("STEP 1: Load reference data")
    categories = load_categories(source_dir / 'CategoriesFC.xlsx')
    
    # Step 2: Load main data
    print("\nSTEP 2: Load main historical data")
    df = load_foodcost_data(source_dir / 'FoodCost.xlsx')
    
    # Step 3: Optional cross-validation
    print("\nSTEP 3: Load ΚΟΣΤΟΛΟΓΗΣΗ for cross-check (optional)")
    kostologisi = {}
    kosto_files = list(source_dir.glob('*ΚΟΣΤΟΛΟΓΗΣΗ*.xlsx'))
    if kosto_files:
        kostologisi = load_kostologisi(kosto_files[0])
    else:
        print("    → No ΚΟΣΤΟΛΟΓΗΣΗ file found, skipping")
    
    # Step 3b: Load Μεικτό κέρδος (product profitability)
    print("\nSTEP 3b: Load Μεικτό κέρδος (optional)")
    meikto_data = None
    # Match all variations of Μεικτό κέρδος filename (with space, underscore, accents)
    meikto_files = (
        list(source_dir.glob('*Μεικτό_κέρδος*.xlsx'))
        + list(source_dir.glob('*Μεικτό κέρδος*.xlsx'))
        + list(source_dir.glob('*Μεικτο_κερδος*.xlsx'))
        + list(source_dir.glob('*Μεικτο κερδος*.xlsx'))
    )
    # Deduplicate (in case multiple patterns match same file)
    meikto_files = list(set(meikto_files))
    if meikto_files:
        meikto_data = load_meikto_kerdos(meikto_files[0])
    else:
        print("    → No Μεικτό κέρδος file found, skipping (Products tab will be empty)")
    
    # Step 3c: Load ALC + Combos hourly
    print("\nSTEP 3c: Load ALC + Combos hourly (optional)")
    alc_data = None
    combos_data = None
    alc_files = list(source_dir.glob('*ALC*Stores*Hours*.xlsx'))
    combos_files = list(source_dir.glob('*Combos*Stores*Hours*.xlsx'))
    if alc_files:
        try:
            alc_data = load_alc_combos(alc_files[0], 'alc')
        except Exception as e:
            print(f"    → ALC load failed: {e}")
    else:
        print("    → No ALC file found, skipping")
    if combos_files:
        try:
            combos_data = load_alc_combos(combos_files[0], 'combos')
        except Exception as e:
            print(f"    → Combos load failed: {e}")
    else:
        print("    → No Combos file found, skipping")
    
    # Step 4: Aggregate
    print("\nSTEP 4: Aggregate monthly")
    monthly_records = aggregate_monthly(df, categories)
    
    # Step 5: Top variance ingredients for latest month
    print("\nSTEP 5: Compute top-variance ingredients (latest month)")
    latest_year = int(df['year'].max())
    latest_month = int(df[df['year'] == latest_year]['month'].max())
    print(f"    Latest period: {latest_year}-{latest_month:02d}")
    top_variances = aggregate_top_variances(df, categories, latest_year, latest_month)
    
    # Step 6: Build final JSON
    print("\nSTEP 6: Build JSON output")
    
    # Collect available periods
    periods = sorted({(r['year'], r['month']) for r in monthly_records})
    period_strings = [f"{y}-{m:02d}" for y, m in periods]
    
    # Build categories list with totals
    all_categories = sorted(set(categories.values()))
    
    # ========== REALTIME + DAILY ARCHIVE ==========
    # Load Quantity Per Item (waste in quantities) from FoodCost.xlsx
    print("\nSTEP 6.5: Load 'Quantity Per Item' sheet (waste quantities)")
    foodcost_path = source_dir / 'FoodCost.xlsx'
    qty_per_item = load_quantity_per_item(foodcost_path) if foodcost_path.exists() else None
    
    print("\nSTEP 7: Compute real-time metrics + daily archive")
    realtime = compute_realtime_metrics(meikto_data, monthly_records, KFC_STORES) if meikto_data else None
    if realtime:
        print(f"  → Real-time chain: €{realtime['chain']['sales']:,.0f} sales, FC% {realtime['chain']['fc_pct']*100:.2f}%")
        print(f"  → Per-store: {len(realtime['per_store'])} stores")
        print(f"  → Hourly: {len(realtime['hourly'])} hours")
        print(f"  → Top FC contributors: {len(realtime['top_fc'])} products")
    
    # Daily archive (saves today's snapshot, returns full history)
    archive_dir = output_dir / 'daily_archive'
    daily_history = update_daily_archive(realtime, archive_dir)
    
    output = {
        'meta': {
            'generated_at': datetime.now(timezone.utc).isoformat(timespec='seconds'),
            'source_files': [f for f in [
                'FoodCost.xlsx', 'CategoriesFC.xlsx',
                next((f.name for f in kosto_files), None),
                next((f.name for f in meikto_files), None) if meikto_files else None,
                next((f.name for f in alc_files), None) if alc_files else None,
                next((f.name for f in combos_files), None) if combos_files else None,
            ] if f],
            'kfc_stores': [STORE_NAME_MAP[s] for s in KFC_STORES],
            'periods': period_strings,
            'latest_period': f"{latest_year}-{latest_month:02d}",
            'categories': all_categories,
            'has_products': meikto_data is not None,
            'has_hourly_alc': alc_data is not None,
            'has_hourly_combos': combos_data is not None,
        },
        'monthly': monthly_records,
        'top_variances_latest': top_variances,
        'kostologisi': [
            {'year': y, 'month': m, 'store_raw': s, 'cogs': v['cogs']}
            for (y, m, s), v in kostologisi.items()
        ] if kostologisi else [],
        'product_profit': meikto_data['products'] if meikto_data else [],
        'store_totals': meikto_data['store_totals'] if meikto_data else [],
        'store_hourly': meikto_data['store_hourly'] if meikto_data else {},
        'store_salestype': meikto_data['store_salestype'] if meikto_data else {},
        'hourly_category': meikto_data['hourly_category'] if meikto_data else {},
        'hourly_alc': alc_data['stores'] if alc_data else {},
        'hourly_combos': combos_data['stores'] if combos_data else {},
        'realtime': realtime if realtime else None,
        'qty_per_item': qty_per_item,
        'daily_history': daily_history if daily_history else [],
    }
    
    # Write
    output_path = output_dir / 'food_data.json'
    output_dir.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=None, separators=(',',':'))
    
    size_kb = output_path.stat().st_size / 1024
    print(f"\n✓ Wrote {output_path.name} ({size_kb:.1f} KB)")
    print(f"  - {len(monthly_records)} monthly records")
    print(f"  - {len(periods)} periods covered: {period_strings[0]} → {period_strings[-1]}")
    print(f"  - {len(all_categories)} categories")
    print(f"  - {len(top_variances)} stores with variance breakdown for {latest_year}-{latest_month:02d}")
    
    return output


# =====================================================================
# CLI
# =====================================================================
def main():
    parser = argparse.ArgumentParser(description='Build food_data.json for the Food Cost Hub')
    parser.add_argument('--source-dir', default='/mnt/user-data/uploads',
                        help='Folder containing source xlsx files')
    parser.add_argument('--output-dir', default='/mnt/user-data/outputs/food-cost-hub',
                        help='Output folder for food_data.json')
    args = parser.parse_args()
    
    source = Path(args.source_dir)
    output = Path(args.output_dir)
    
    if not source.exists():
        print(f"ERROR: Source directory not found: {source}", file=sys.stderr)
        sys.exit(1)
    
    build_food_data_json(source, output)


if __name__ == '__main__':
    main()
